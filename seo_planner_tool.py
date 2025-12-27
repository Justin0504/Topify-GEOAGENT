"""
title: SEO 规划工具
description: 【关键词研究】调用API获取真实关键词数据并生成Excel | 【页面映射】生成页面-关键词映射表 | 【内容规划】生成博客文章SEO内容计划
author: GEO Agent
version: 2.0.0
required_open_webui_version: 0.6.0
requirements: openpyxl, requests
"""

import os
import io
import json
import requests
import re
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
from pydantic import BaseModel, Field

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class Tools:
    """
    SEO 规划工具 - 关键词研究、页面映射、内容规划（调用真实 API 数据）
    
    ═══════════════════════════════════════════════════════════════
    🎯 快速匹配指南
    ═══════════════════════════════════════════════════════════════
    
    🔍 用户说"关键词研究"、"扩展关键词"、"SEO关键词"
       → 调用 keyword_research（会调用 API 获取真实数据）
    
    📄 用户说"页面映射"、"关键词映射"、"页面优化哪些词"
       → 调用 page_keyword_mapping
    
    📅 用户说"内容规划"、"博客计划"、"文章规划"、"Pillar文章"
       → 调用 content_planning
    
    ═══════════════════════════════════════════════════════════════
    
    **注意**: 需要在 Valves 中配置 API_KEY 才能获取真实 SEO 数据
    """

    class Valves(BaseModel):
        OUTPUT_PATH: str = Field(
            default="/app/backend/data/output",
            description="文件保存路径"
        )
        API_KEY: str = Field(
            default="",
            description="【必填】SEO 分析 API Key（Semrush）"
        )
        DEFAULT_DATABASE: str = Field(
            default="us",
            description="默认数据库/地区代码 (us=美国, cn=中国, uk=英国)"
        )
        DEFAULT_LIMIT: int = Field(
            default=50,
            description="默认返回关键词数量"
        )

    def __init__(self):
        self.valves = self.Valves()
        self.api_base_url = "https://api.semrush.com/"

    # ==================== API 调用方法 ====================
    
    def _make_api_request(self, params: dict) -> dict:
        """
        发送 SEO API 请求
        
        :param params: API 参数
        :return: 解析后的响应数据
        """
        api_key = self.valves.API_KEY.strip()
        
        if not api_key:
            return {
                "success": False,
                "error": "未配置 API Key，将使用模板数据",
                "data": []
            }
        
        params["key"] = api_key
        
        try:
            response = requests.get(self.api_base_url, params=params, timeout=30)
            raw_text = response.text.strip()
            
            # 检查错误响应
            if raw_text.startswith("ERROR"):
                error_code = raw_text.split("::")[0] if "::" in raw_text else raw_text
                error_messages = {
                    "ERROR 50": "API Key 无效或已过期",
                    "ERROR 40": "超出 API 调用限制",
                    "ERROR 120": "无效的数据库/地区代码",
                    "ERROR 130": "请求的数据库中没有此数据"
                }
                parts = error_code.split(" ")
                key = f"{parts[0]} {parts[1]}" if len(parts) > 1 else error_code
                friendly_error = error_messages.get(key, raw_text)
                return {"success": False, "error": f"API 错误: {friendly_error}", "data": []}
            
            # 解析 CSV 格式响应（Semrush 使用分号分隔）
            lines = raw_text.split("\n")
            if not lines or not lines[0]:
                return {"success": True, "data": [], "columns": [], "count": 0}
            
            # 获取列名（第一行）
            columns = lines[0].split(";")
            
            # 解析数据行
            data = []
            for line in lines[1:]:
                if line.strip():
                    values = line.split(";")
                    # 确保值的数量和列名匹配
                    if len(values) >= len(columns):
                        row = dict(zip(columns, values))
                        data.append(row)
                    elif len(values) > 0:
                        # 如果值少于列名，用空字符串填充
                        padded_values = values + [''] * (len(columns) - len(values))
                        row = dict(zip(columns, padded_values))
                        data.append(row)
            
            return {
                "success": True,
                "data": data,
                "columns": columns,
                "count": len(data),
                "raw_sample": raw_text[:500] if raw_text else ""  # 保存原始响应样本用于调试
            }
            
        except requests.exceptions.Timeout:
            return {"success": False, "error": "请求超时", "data": []}
        except requests.exceptions.RequestException as e:
            return {"success": False, "error": f"网络错误: {str(e)}", "data": []}
        except Exception as e:
            return {"success": False, "error": f"解析错误: {str(e)}", "data": []}

    def _get_domain_keywords(self, domain: str, limit: int = 50, database: str = "us") -> tuple:
        """获取域名的排名关键词，返回 (data, error_msg, debug_info)"""
        params = {
            "type": "domain_organic",
            "domain": domain,
            "database": database,
            "display_limit": limit,
            "display_sort": "tr_desc",
            "export_columns": "Ph,Po,Nq,Cp,Co,Kd,Tr,Ur"
        }
        result = self._make_api_request(params)
        if not result.get("success"):
            return [], result.get("error", "未知错误"), None
        
        # 调试信息：返回列名和第一条数据
        debug_info = {
            "columns": result.get("columns", []),
            "sample": result.get("data", [])[0] if result.get("data") else {},
            "raw_sample": result.get("raw_sample", "")[:200]
        }
        return result.get("data", []), None, debug_info

    def _get_related_keywords(self, keyword: str, limit: int = 30, database: str = "us") -> tuple:
        """获取相关关键词，返回 (data, error_msg, debug_info)"""
        params = {
            "type": "phrase_related",
            "phrase": keyword,
            "database": database,
            "display_limit": limit,
            "export_columns": "Ph,Nq,Cp,Co,Kd,Nr"
        }
        result = self._make_api_request(params)
        if not result.get("success"):
            return [], result.get("error", "未知错误"), None
        
        debug_info = {
            "columns": result.get("columns", []),
            "sample": result.get("data", [])[0] if result.get("data") else {},
        }
        return result.get("data", []), None, debug_info

    def _get_question_keywords(self, keyword: str, limit: int = 20, database: str = "us") -> tuple:
        """获取问题类关键词，返回 (data, error_msg, debug_info)"""
        params = {
            "type": "phrase_questions",
            "phrase": keyword,
            "database": database,
            "display_limit": limit,
            "export_columns": "Ph,Nq,Cp,Co,Kd"
        }
        result = self._make_api_request(params)
        if not result.get("success"):
            return [], result.get("error", "未知错误"), None
        
        debug_info = {
            "columns": result.get("columns", []),
            "sample": result.get("data", [])[0] if result.get("data") else {},
        }
        return result.get("data", []), None, debug_info

    def _get_competitors(self, domain: str, limit: int = 10, database: str = "us") -> list:
        """获取竞争对手"""
        params = {
            "type": "domain_organic_organic",
            "domain": domain,
            "database": database,
            "display_limit": limit,
            "export_columns": "Dn,Cr,Np,Or,Ot"
        }
        result = self._make_api_request(params)
        return result.get("data", [])

    # ==================== Excel 辅助方法 ====================

    def _save_excel(self, wb, filename: str) -> str:
        """保存 Excel 文件"""
        output_dir = self.valves.OUTPUT_PATH
        try:
            os.makedirs(output_dir, exist_ok=True)
            file_path = os.path.join(output_dir, filename)
            wb.save(file_path)
            
            if os.path.exists(file_path):
                size = os.path.getsize(file_path)
                size_str = f"{size / 1024:.1f} KB" if size >= 1024 else f"{size} bytes"
                return f"""✅ Excel 文件已保存！

📄 文件名: {filename}
📁 路径: {file_path}
📊 大小: {size_str}
🕐 时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
"""
            else:
                return f"❌ 文件保存失败"
        except Exception as e:
            return f"❌ 保存失败: {str(e)}"

    def _apply_header_style(self, ws, row: int, col_count: int):
        """应用表头样式"""
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _apply_data_style(self, ws, start_row: int, end_row: int, col_count: int):
        """应用数据行样式"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    # ==================== 主要功能方法 ====================

    def keyword_research(
        self,
        website_url: str,
        product_services: str,
        target_market: str = "美国",
        language: str = "英文",
        keyword_count: int = 50,
        __user__: dict = None
    ) -> str:
        """
        🔍 关键词研究 - 调用 API 获取真实关键词数据并生成 Excel（REQ-001）
        
        直接使用 Semrush API 返回的原始 CSV 格式，不做转换。
        
        :param website_url: 【必填】网站URL
        :param product_services: 【必填】产品/服务描述
        :param target_market: 目标市场（默认：美国）
        :param language: 关键词语言（默认：英文）
        :param keyword_count: 生成关键词数量（默认：50）
        :return: 包含真实数据的关键词列表 Excel 文件
        """
        if not website_url or not product_services:
            return "❌ 请提供网站URL和产品/服务描述"
        
        # 清理 URL
        website_url = website_url.strip().lower()
        domain = website_url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
        
        # 地区映射
        market_to_db = {
            "美国": "us", "中国": "cn", "英国": "uk", "德国": "de", 
            "法国": "fr", "日本": "jp", "韩国": "kr", "加拿大": "ca",
            "澳大利亚": "au", "新加坡": "sg", "香港": "hk"
        }
        database = market_to_db.get(target_market, "us")
        
        # 检查 API Key
        has_api = bool(self.valves.API_KEY.strip())
        if not has_api:
            return "❌ 未配置 API Key，请在工具 Valves 中配置 API_KEY"
        
        # 创建工作簿
        wb = Workbook()
        
        api_errors = []
        debug_info_list = []
        total_keywords = 0
        
        # ==================== Sheet 1: 域名关键词（原始格式）====================
        
        domain_keywords, domain_error, domain_debug = self._get_domain_keywords(
            domain, limit=keyword_count, database=database
        )
        
        if domain_error:
            api_errors.append(f"域名关键词: {domain_error}")
        if domain_debug:
            debug_info_list.append(f"域名API列名: {domain_debug.get('columns')}")
        
        ws1 = wb.active
        ws1.title = "域名关键词"
        
        if domain_keywords and domain_debug:
            # 使用 API 返回的原始列名
            columns = domain_debug.get("columns", [])
            
            # 写入表头（原始列名）
            for col, header in enumerate(columns, 1):
                ws1.cell(row=1, column=col, value=header)
                ws1.cell(row=1, column=col).font = Font(bold=True)
                ws1.cell(row=1, column=col).fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
                ws1.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            
            # 写入数据（按原始列顺序）
            for row_idx, kw_data in enumerate(domain_keywords, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=kw_data.get(col_name, ""))
                total_keywords += 1
            
            # 调整列宽
            for col in range(1, len(columns) + 1):
                ws1.column_dimensions[get_column_letter(col)].width = 15
        else:
            ws1.cell(row=1, column=1, value="无数据（该域名在 Semrush 数据库中可能没有记录）")
        
        # ==================== Sheet 2: 相关关键词（原始格式）====================
        
        # 从产品描述中提取核心关键词
        english_words = re.findall(r'[A-Za-z]+(?:\s+[A-Za-z]+)*', product_services)
        if english_words:
            core_keywords = max(english_words, key=len)[:50]
        else:
            core_keywords = product_services.split("，")[0].split(",")[0][:30]
        
        related_keywords, related_error, related_debug = self._get_related_keywords(
            core_keywords, limit=keyword_count, database=database
        )
        
        if related_error:
            api_errors.append(f"相关关键词({core_keywords}): {related_error}")
        if related_debug:
            debug_info_list.append(f"相关词API列名: {related_debug.get('columns')}")
        
        ws2 = wb.create_sheet("相关关键词")
        ws2.cell(row=1, column=1, value=f"搜索词: {core_keywords}")
        ws2.cell(row=1, column=1).font = Font(bold=True, color="2c3e50")
        
        if related_keywords and related_debug:
            columns = related_debug.get("columns", [])
            
            # 写入表头
            for col, header in enumerate(columns, 1):
                ws2.cell(row=2, column=col, value=header)
                ws2.cell(row=2, column=col).font = Font(bold=True, color="FFFFFF")
                ws2.cell(row=2, column=col).fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            
            # 写入数据
            for row_idx, kw_data in enumerate(related_keywords, 3):
                for col_idx, col_name in enumerate(columns, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=kw_data.get(col_name, ""))
                total_keywords += 1
            
            for col in range(1, len(columns) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 15
        else:
            ws2.cell(row=2, column=1, value="无数据")
        
        # ==================== Sheet 3: 问题关键词（原始格式）====================
        
        question_keywords, question_error, question_debug = self._get_question_keywords(
            core_keywords, limit=keyword_count // 2, database=database
        )
        
        if question_error:
            api_errors.append(f"问题关键词: {question_error}")
        if question_debug:
            debug_info_list.append(f"问题词API列名: {question_debug.get('columns')}")
        
        ws3 = wb.create_sheet("问题关键词")
        ws3.cell(row=1, column=1, value=f"搜索词: {core_keywords}")
        ws3.cell(row=1, column=1).font = Font(bold=True, color="2c3e50")
        
        if question_keywords and question_debug:
            columns = question_debug.get("columns", [])
            
            # 写入表头
            for col, header in enumerate(columns, 1):
                ws3.cell(row=2, column=col, value=header)
                ws3.cell(row=2, column=col).font = Font(bold=True, color="FFFFFF")
                ws3.cell(row=2, column=col).fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            
            # 写入数据
            for row_idx, kw_data in enumerate(question_keywords, 3):
                for col_idx, col_name in enumerate(columns, 1):
                    ws3.cell(row=row_idx, column=col_idx, value=kw_data.get(col_name, ""))
                total_keywords += 1
            
            for col in range(1, len(columns) + 1):
                ws3.column_dimensions[get_column_letter(col)].width = 20
        else:
            ws3.cell(row=2, column=1, value="无数据")
        
        # ==================== Sheet 4: 说明 ====================
        
        ws4 = wb.create_sheet("说明")
        ws4.cell(row=1, column=1, value="Semrush API 列名说明")
        ws4.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        column_explanations = [
            ("Ph", "Phrase - 关键词"),
            ("Nq", "Number of Queries - 月搜索量"),
            ("Cp", "CPC - 每次点击成本（美元）"),
            ("Co", "Competition - 竞争度（0-1）"),
            ("Kd", "Keyword Difficulty - SEO难度（0-100）"),
            ("Nr", "Number of Results - 搜索结果数"),
            ("Td", "Trend - 趋势"),
            ("Po", "Position - 排名位置"),
            ("Pp", "Previous Position - 上次排名"),
            ("Pd", "Position Difference - 排名变化"),
            ("Tr", "Traffic - 流量"),
            ("Tc", "Traffic Cost - 流量成本"),
            ("Ur", "URL - 排名页面URL"),
        ]
        
        for i, (code, desc) in enumerate(column_explanations, 3):
            ws4.cell(row=i, column=1, value=code)
            ws4.cell(row=i, column=1).font = Font(bold=True)
            ws4.cell(row=i, column=2, value=desc)
        
        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 40
        
        # 保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"keyword_research_{domain.replace('.', '_')}_{timestamp}.xlsx"
        
        save_result = self._save_excel(wb, filename)
        
        # 生成报告
        api_status = ""
        if total_keywords > 0:
            api_status = "✅ API 调用成功"
        elif api_errors:
            api_status = "⚠️ API 调用遇到问题:\n" + "\n".join([f"  • {e}" for e in api_errors])
        else:
            api_status = "⚠️ API 未返回数据"
        
        return save_result + f"""
📊 **关键词研究报告已生成（原始 Semrush CSV 格式）**

🌐 网站: {domain}
🎯 目标市场: {target_market} ({database.upper()})
🔑 核心搜索词: {core_keywords}
📈 关键词总数: {total_keywords}

🔧 **API 状态**: {api_status}

📋 **Excel 工作表**:
1. 域名关键词 - {domain} 当前排名的关键词
2. 相关关键词 - 与 "{core_keywords}" 相关的词
3. 问题关键词 - 问题类长尾词
4. 说明 - Semrush 列名解释

{chr(10).join(debug_info_list) if debug_info_list else ""}
"""

    def page_keyword_mapping(
        self,
        website_url: str,
        limit: int = 100,
        __user__: dict = None
    ) -> str:
        """
        📄 页面-关键词映射 - 获取网站每个页面当前排名的关键词（真实数据）
        
        ════════════════════════════════════════════════════════
        🎯 何时使用此工具
        ════════════════════════════════════════════════════════
        
        ✅ "生成页面关键词映射"、"页面排名了哪些词"
        ✅ "查看每个页面的关键词"、"页面SEO分析"
        ✅ "关键词映射表"、"URL关键词分析"
        
        ════════════════════════════════════════════════════════
        📋 参数说明
        ════════════════════════════════════════════════════════
        
        :param website_url: 【必填】网站URL，如 "topify.ai"
        :param limit: 获取关键词数量（默认 100）
        :return: 按页面分组的关键词映射 Excel 文件
        
        ════════════════════════════════════════════════════════
        📊 功能说明
        ════════════════════════════════════════════════════════
        
        调用 Semrush API 获取网站当前排名数据，显示：
        - 每个页面（URL）排名了哪些关键词
        - 每个关键词的排名位置、搜索量、流量等
        - 按 URL 分组，方便查看每个页面的 SEO 表现
        """
        if not website_url:
            return "❌ 请提供网站URL"
        
        domain = website_url.strip().lower()
        domain = domain.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
        
        # 检查 API Key
        has_api = bool(self.valves.API_KEY.strip())
        if not has_api:
            return "❌ 未配置 API Key，请在工具 Valves 中配置 API_KEY"
        
        # 获取域名关键词数据（包含 URL）
        api_keywords, api_error, api_debug = self._get_domain_keywords(
            domain, limit=limit, database=self.valves.DEFAULT_DATABASE
        )
        
        if api_error:
            return f"❌ API 错误: {api_error}"
        
        if not api_keywords:
            return f"❌ {domain} 在 Semrush 数据库中没有排名数据"
        
        # 按 URL 分组
        url_keywords = {}
        for kw in api_keywords:
            url = kw.get("Ur", "未知页面")
            if url not in url_keywords:
                url_keywords[url] = []
            url_keywords[url].append(kw)
        
        # 创建工作簿
        wb = Workbook()
        
        # ==================== Sheet 1: 按页面分组的汇总 ====================
        ws1 = wb.active
        ws1.title = "页面关键词汇总"
        
        # 表头
        headers = ["页面URL", "关键词数量", "总流量", "最高排名关键词", "最高排名位置"]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
            ws1.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            ws1.cell(row=1, column=col).fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        
        row = 2
        for url, keywords in sorted(url_keywords.items(), key=lambda x: len(x[1]), reverse=True):
            # 计算总流量
            total_traffic = 0
            best_keyword = ""
            best_position = 999
            
            for kw in keywords:
                try:
                    traffic = float(kw.get("Tr", 0) or 0)
                    total_traffic += traffic
                except:
                    pass
                
                try:
                    pos = int(kw.get("Po", 999) or 999)
                    if pos < best_position:
                        best_position = pos
                        best_keyword = kw.get("Ph", "")
                except:
                    pass
            
            ws1.cell(row=row, column=1, value=url)
            ws1.cell(row=row, column=2, value=len(keywords))
            ws1.cell(row=row, column=3, value=round(total_traffic, 1))
            ws1.cell(row=row, column=4, value=best_keyword)
            ws1.cell(row=row, column=5, value=best_position if best_position < 999 else "N/A")
            row += 1
        
        # 调整列宽
        ws1.column_dimensions['A'].width = 50
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 15
        
        # ==================== Sheet 2: 原始数据（完整 API 返回）====================
        ws2 = wb.create_sheet("原始数据")
        
        if api_debug:
            columns = api_debug.get("columns", [])
            
            # 写入表头
            for col, header in enumerate(columns, 1):
                ws2.cell(row=1, column=col, value=header)
                ws2.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
                ws2.cell(row=1, column=col).fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            
            # 写入数据
            for row_idx, kw_data in enumerate(api_keywords, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=kw_data.get(col_name, ""))
            
            # 调整列宽
            for col in range(1, len(columns) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 15
        
        # ==================== Sheet 3: 按页面分组的详细数据 ====================
        ws3 = wb.create_sheet("按页面分组")
        
        row = 1
        for url, keywords in sorted(url_keywords.items(), key=lambda x: len(x[1]), reverse=True):
            # 页面标题
            ws3.cell(row=row, column=1, value=f"📄 {url}")
            ws3.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws3.cell(row=row, column=1).fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            ws3.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # 小表头
            sub_headers = ["关键词(Ph)", "排名(Po)", "搜索量(Nq)", "流量(Tr)", "CPC(Cp)", "难度(Kd)"]
            for col, header in enumerate(sub_headers, 1):
                ws3.cell(row=row, column=col, value=header)
                ws3.cell(row=row, column=col).font = Font(bold=True)
                ws3.cell(row=row, column=col).fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            row += 1
            
            # 关键词数据
            for kw in sorted(keywords, key=lambda x: int(x.get("Po", 999) or 999)):
                ws3.cell(row=row, column=1, value=kw.get("Ph", ""))
                ws3.cell(row=row, column=2, value=kw.get("Po", ""))
                ws3.cell(row=row, column=3, value=kw.get("Nq", ""))
                ws3.cell(row=row, column=4, value=kw.get("Tr", ""))
                ws3.cell(row=row, column=5, value=kw.get("Cp", ""))
                ws3.cell(row=row, column=6, value=kw.get("Kd", ""))
                row += 1
            
            # 空行分隔
            row += 1
        
        # 调整列宽
        ws3.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 12
        
        # ==================== Sheet 4: 说明 ====================
        ws4 = wb.create_sheet("说明")
        ws4.cell(row=1, column=1, value="Semrush API 列名说明")
        ws4.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        column_explanations = [
            ("Ph", "Phrase - 关键词"),
            ("Po", "Position - 当前排名位置"),
            ("Pp", "Previous Position - 上月排名"),
            ("Pd", "Position Difference - 排名变化"),
            ("Nq", "Number of Queries - 月搜索量"),
            ("Cp", "CPC - 每次点击成本（美元）"),
            ("Co", "Competition - 竞争度（0-1）"),
            ("Kd", "Keyword Difficulty - SEO难度（0-100）"),
            ("Tr", "Traffic - 预估流量"),
            ("Tc", "Traffic Cost - 流量价值（美元）"),
            ("Ur", "URL - 排名页面地址"),
        ]
        
        for i, (code, desc) in enumerate(column_explanations, 3):
            ws4.cell(row=i, column=1, value=code)
            ws4.cell(row=i, column=1).font = Font(bold=True)
            ws4.cell(row=i, column=2, value=desc)
        
        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 40
        
        # 保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"page_keyword_mapping_{domain.replace('.', '_')}_{timestamp}.xlsx"
        
        save_result = self._save_excel(wb, filename)
        
        # 统计信息
        total_keywords = len(api_keywords)
        total_pages = len(url_keywords)
        
        return save_result + f"""
📄 **页面-关键词映射表已生成（Semrush 真实数据）**

🌐 网站: {domain}
📊 数据库: {self.valves.DEFAULT_DATABASE.upper()}
📈 总关键词数: {total_keywords}
📄 页面数量: {total_pages}

📋 **Excel 工作表**:
1. **页面关键词汇总** - 每个页面的关键词数量、流量、最佳排名
2. **原始数据** - Semrush API 完整返回数据
3. **按页面分组** - 每个 URL 排名的关键词详情
4. **说明** - 列名解释

🔝 **排名最多关键词的页面**:
{chr(10).join([f"  • {url[:50]}... ({len(kws)}个词)" for url, kws in sorted(url_keywords.items(), key=lambda x: len(x[1]), reverse=True)[:5]])}
"""

    def content_planning(
        self,
        website_url: str,
        product_services: str,
        pillar_topics: str,
        months: int = 6,
        articles_per_month: int = 13,
        __user__: dict = None
    ) -> str:
        """
        📅 内容规划 - 生成博客SEO内容计划（REQ-003）
        
        ════════════════════════════════════════════════════════
        🎯 何时使用此工具
        ════════════════════════════════════════════════════════
        
        ✅ "生成内容规划"、"博客文章计划"、"内容日历"
        ✅ "Pillar文章规划"、"支柱内容规划"
        ✅ "6个月内容计划"、"SEO文章规划"
        
        ════════════════════════════════════════════════════════
        📋 参数说明
        ════════════════════════════════════════════════════════
        
        :param website_url: 【必填】网站URL
        :param product_services: 【必填】产品/服务描述
        :param pillar_topics: 【必填】支柱主题，用逗号分隔
            示例: "GEO优化指南, AI搜索营销, 品牌AI可见性"
        :param months: 规划月数（默认：6）
        :param articles_per_month: 每月文章数（默认：13，约80篇/6个月）
        :return: 内容规划 Excel 文件
        
        ════════════════════════════════════════════════════════
        📊 Pillar-Based Marketing 策略
        ════════════════════════════════════════════════════════
        
        支柱文章（Pillar Content）：
        - 全面、权威的长篇内容（3000-5000字）
        - 覆盖某个主题的方方面面
        - 内链到多篇支撑文章
        
        支撑文章（Cluster Content）：
        - 针对具体子话题的文章（1000-2000字）
        - 回链到对应的支柱文章
        - 形成主题簇，增强SEO权重
        """
        if not website_url or not product_services or not pillar_topics:
            return "❌ 请提供网站URL、产品描述和支柱主题"
        
        pillar_list = [p.strip() for p in re.split(r'[,;]+', pillar_topics) if p.strip()]
        domain = website_url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
        
        total_articles = months * articles_per_month
        
        # 获取问题类关键词作为文章灵感
        has_api = bool(self.valves.API_KEY.strip())
        question_keywords = []
        
        if has_api and pillar_list:
            # 从第一个支柱主题获取问题关键词
            question_keywords, _, _ = self._get_question_keywords(
                pillar_list[0], 
                limit=20, 
                database=self.valves.DEFAULT_DATABASE
            )
        
        # 创建工作簿
        wb = Workbook()
        
        # === Sheet 1: 内容规划总览 ===
        ws1 = wb.active
        ws1.title = "内容规划总览"
        
        ws1.merge_cells('A1:I1')
        title_cell = ws1['A1']
        title_cell.value = f"SEO 内容规划 - {domain}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="8e44ad", end_color="8e44ad", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 35
        
        # 规划概要
        ws1['A3'] = "📊 规划概要"
        ws1['A3'].font = Font(size=12, bold=True)
        ws1['A4'] = f"规划周期: {months} 个月"
        ws1['A5'] = f"总文章数: {total_articles} 篇"
        ws1['A6'] = f"支柱主题: {len(pillar_list)} 个"
        ws1['A7'] = f"每月文章: {articles_per_month} 篇"
        
        # 表头
        headers = ["序号", "发布日期", "文章类型", "支柱主题", "文章标题", "目标关键词", "字数要求", "状态", "负责人"]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=9, column=col, value=header)
        self._apply_header_style(ws1, 9, len(headers))
        
        # 生成内容计划
        row = 10
        start_date = datetime.now()
        article_idx = 1
        
        pillar_fill = PatternFill(start_color="f5b7b1", end_color="f5b7b1", fill_type="solid")
        
        for month in range(months):
            month_date = start_date + timedelta(days=month * 30)
            
            for week in range(4):
                for article in range(articles_per_month // 4 + (1 if week < articles_per_month % 4 else 0)):
                    if article_idx > total_articles:
                        break
                    
                    pub_date = month_date + timedelta(days=week * 7 + article * 2)
                    pillar_topic = pillar_list[article_idx % len(pillar_list)]
                    
                    # 每个支柱主题第一篇是支柱文章，其余是支撑文章
                    is_pillar = (article_idx <= len(pillar_list)) or (article_idx % 10 == 1)
                    article_type = "🏛️ 支柱文章" if is_pillar else "📄 支撑文章"
                    word_count = "3000-5000" if is_pillar else "1000-2000"
                    
                    # 如果有问题关键词，用作文章标题建议
                    suggested_title = ""
                    if question_keywords and article_idx <= len(question_keywords):
                        suggested_title = question_keywords[article_idx - 1].get("Ph", "")
                    
                    ws1.cell(row=row, column=1, value=article_idx)
                    ws1.cell(row=row, column=2, value=pub_date.strftime("%Y-%m-%d"))
                    ws1.cell(row=row, column=3, value=article_type)
                    ws1.cell(row=row, column=4, value=pillar_topic)
                    ws1.cell(row=row, column=5, value=suggested_title)  # 标题建议
                    ws1.cell(row=row, column=6, value="")  # 关键词待填
                    ws1.cell(row=row, column=7, value=word_count)
                    ws1.cell(row=row, column=8, value="待撰写")
                    ws1.cell(row=row, column=9, value="")
                    
                    # 高亮支柱文章
                    if is_pillar:
                        for col in range(1, 10):
                            ws1.cell(row=row, column=col).fill = pillar_fill
                    
                    article_idx += 1
                    row += 1
        
        # 调整列宽
        col_widths = [8, 12, 14, 20, 40, 25, 12, 10, 10]
        for i, width in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = width
        
        # === Sheet 2: 支柱主题详情 ===
        ws2 = wb.create_sheet("支柱主题详情")
        
        ws2['A1'] = "支柱主题 (Pillar Topics)"
        ws2['A1'].font = Font(size=14, bold=True)
        
        ws2['A3'] = "支柱主题"
        ws2['B3'] = "支柱文章标题"
        ws2['C3'] = "支撑文章数量"
        ws2['D3'] = "核心关键词"
        ws2['E3'] = "内链策略"
        
        for col in range(1, 6):
            ws2.cell(row=3, column=col).font = Font(bold=True)
            ws2.cell(row=3, column=col).fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            ws2.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
        
        for i, pillar in enumerate(pillar_list, 4):
            ws2.cell(row=i, column=1, value=pillar)
            ws2.cell(row=i, column=2, value="")  # 待填
            ws2.cell(row=i, column=3, value=f"~{total_articles // len(pillar_list) - 1} 篇")
            ws2.cell(row=i, column=4, value="")  # 待填
            ws2.cell(row=i, column=5, value="支撑文章回链到此支柱文章")
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 30
        
        # === Sheet 3: 关键词灵感（如果有API数据）===
        if question_keywords:
            ws3 = wb.create_sheet("关键词灵感")
            
            ws3['A1'] = "问题类关键词灵感（来自 API）"
            ws3['A1'].font = Font(size=14, bold=True)
            
            ws3['A3'] = "关键词"
            ws3['B3'] = "月搜索量"
            ws3['C3'] = "SEO难度"
            ws3['D3'] = "建议用途"
            
            for col in range(1, 5):
                ws3.cell(row=3, column=col).font = Font(bold=True)
            
            for i, kw_data in enumerate(question_keywords, 4):
                ws3.cell(row=i, column=1, value=kw_data.get("Ph", ""))
                ws3.cell(row=i, column=2, value=kw_data.get("Nq", ""))
                ws3.cell(row=i, column=3, value=kw_data.get("Kd", ""))
                ws3.cell(row=i, column=4, value="博客文章标题")
            
            ws3.column_dimensions['A'].width = 50
            ws3.column_dimensions['B'].width = 12
            ws3.column_dimensions['C'].width = 10
            ws3.column_dimensions['D'].width = 20
        
        # === Sheet 4: 月度统计 ===
        ws4 = wb.create_sheet("月度统计")
        
        ws4['A1'] = "月度发布统计"
        ws4['A1'].font = Font(size=14, bold=True)
        
        ws4['A3'] = "月份"
        ws4['B3'] = "支柱文章"
        ws4['C3'] = "支撑文章"
        ws4['D3'] = "总计"
        ws4['E3'] = "状态"
        
        for col in range(1, 6):
            ws4.cell(row=3, column=col).font = Font(bold=True)
        
        for m in range(months):
            month_date = start_date + timedelta(days=m * 30)
            ws4.cell(row=4+m, column=1, value=month_date.strftime("%Y年%m月"))
            ws4.cell(row=4+m, column=2, value=len(pillar_list) if m == 0 else 1)
            ws4.cell(row=4+m, column=3, value=articles_per_month - (len(pillar_list) if m == 0 else 1))
            ws4.cell(row=4+m, column=4, value=articles_per_month)
            ws4.cell(row=4+m, column=5, value="待开始")
        
        # 保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"content_plan_{domain.replace('.', '_')}_{timestamp}.xlsx"
        
        pillar_count = len(pillar_list) + (total_articles // 10)
        cluster_count = total_articles - pillar_count
        
        return self._save_excel(wb, filename) + f"""
📅 **{months}个月内容规划已生成**

📊 **规划概要**：
• 总文章数: {total_articles} 篇
• 支柱文章: ~{pillar_count} 篇（3000-5000字）
• 支撑文章: ~{cluster_count} 篇（1000-2000字）
• 每月产出: {articles_per_month} 篇

🏛️ **支柱主题**：
{chr(10).join([f"  • {p}" for p in pillar_list])}

{f"💡 **已获取 {len(question_keywords)} 个问题关键词作为文章灵感**" if question_keywords else ""}

📋 **Excel 包含工作表**：
1. 内容规划总览 - 详细的发布日历
2. 支柱主题详情 - 每个支柱的规划
{f"3. 关键词灵感 - API获取的问题关键词" if question_keywords else ""}
{"4" if question_keywords else "3"}. 月度统计 - 进度追踪

💡 **Pillar-Based Marketing 策略**：
1. 支柱文章是某主题的权威长文
2. 支撑文章深入探讨子话题
3. 支撑文章内链到对应支柱文章
4. 形成主题簇，提升整体SEO权重
"""


# ==================== 兼容性别名 ====================
Functions = Tools
Function = Tools
