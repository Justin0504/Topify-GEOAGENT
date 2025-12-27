"""
title: 内容规划工具
description: 【内容规划】生成博客文章SEO规划，Pillar-Based Marketing | 【GEO优化】制定AI回答优化运营计划 | 【项目管理】生成SEO/GEO项目任务清单
author: GEO Agent
version: 2.0.0
required_open_webui_version: 0.6.0
requirements: openpyxl
"""

import os
import re
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
from pydantic import BaseModel, Field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class Tools:
    """
    内容规划工具 - SEO内容规划、GEO优化计划、项目任务管理
    
    ═══════════════════════════════════════════════════════════════
    🎯 功能匹配指南（中文触发词）
    ═══════════════════════════════════════════════════════════════
    
    📅 content_planning - 内容规划
       触发词: "内容规划", "博客计划", "文章规划", "Pillar文章", 
              "支柱内容", "内容日历", "发布计划", "6个月计划"
       示例: "生成6个月80篇文章的SEO内容规划"
       输出: Excel 文件（Pillar-Based 内容计划表）
    
    🤖 geo_optimization_plan - GEO优化计划
       触发词: "GEO优化", "AI优化", "AI回答优化", "GEO计划",
              "AI提示词", "AI搜索优化", "40个提示词"
       示例: "制定GEO运营计划，包含40个AI提示词"
       输出: Excel 文件（AI提示词和监测计划）
    
    📋 project_task_list - 项目任务清单
       触发词: "任务清单", "项目管理", "任务列表", "甘特图",
              "项目计划", "工作清单", "待办事项"
       示例: "生成SEO+GEO项目任务清单"
       输出: Excel 文件（分阶段任务清单）
    
    ═══════════════════════════════════════════════════════════════
    """

    class Valves(BaseModel):
        OUTPUT_PATH: str = Field(
            default="/app/backend/data/output",
            description="文件保存路径（Docker环境）"
        )

    def __init__(self):
        self.valves = self.Valves()

    def _cp_ensure_output_dir(self) -> str:
        """确保输出目录存在"""
        output_path = self.valves.OUTPUT_PATH
        if not os.path.exists(output_path):
            os.makedirs(output_path, exist_ok=True)
        return output_path

    def _cp_style_excel_sheet(self, ws, headers: List[str], header_color: str = "4472C4"):
        """为Excel表格添加样式"""
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.freeze_panes = 'A2'
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # ==================== 辅助方法 ====================
    
    def _cp_generate_seo_url(self, domain: str, keyword: str) -> str:
        """根据关键词生成SEO友好的URL"""
        # 清理关键词，生成URL slug
        slug = keyword.lower().strip()
        # 移除特殊字符，只保留字母数字和空格
        slug = re.sub(r'[^a-z0-9\s-]', '', slug)
        # 替换空格为连字符
        slug = re.sub(r'\s+', '-', slug)
        # 移除多余的连字符
        slug = re.sub(r'-+', '-', slug)
        # 限制长度
        slug = slug[:60].rstrip('-')
        
        # 确保域名格式正确
        if not domain.startswith('http'):
            domain = f"https://{domain}"
        domain = domain.rstrip('/')
        
        return f"{domain}/blog/{slug}"
    
    def _cp_determine_search_intent(self, keyword: str) -> str:
        """根据关键词判断搜索意图"""
        keyword_lower = keyword.lower()
        
        # Transactional 意图
        transactional_words = ['buy', 'purchase', 'price', 'pricing', 'cost', 'discount', 'deal', 'coupon', 'free trial', 'download', 'get', 'order']
        if any(word in keyword_lower for word in transactional_words):
            return "Transactional"
        
        # Commercial 意图
        commercial_words = ['best', 'top', 'review', 'comparison', 'vs', 'versus', 'alternative', 'compare', 'recommend']
        if any(word in keyword_lower for word in commercial_words):
            return "Commercial"
        
        # Navigational 意图
        navigational_words = ['login', 'sign in', 'official', 'website', 'app', 'download']
        if any(word in keyword_lower for word in navigational_words):
            return "Navigational"
        
        # Informational 意图（默认）
        informational_words = ['how', 'what', 'why', 'when', 'where', 'who', 'guide', 'tutorial', 'tips', 'learn', 'example']
        if any(word in keyword_lower for word in informational_words):
            return "Informational"
        
        return "Informational"
    
    def _cp_generate_title_from_keyword(self, keyword: str, intent: str, product_name: str = "") -> str:
        """根据关键词和搜索意图生成文章标题"""
        keyword_clean = keyword.strip()
        
        # 根据意图生成不同风格的标题
        if intent == "Informational":
            templates = [
                f"What is {keyword_clean}? Complete Guide for 2025",
                f"{keyword_clean}: Everything You Need to Know",
                f"The Ultimate Guide to {keyword_clean}",
                f"How to Master {keyword_clean} in 2025",
                f"{keyword_clean} Explained: A Beginner's Guide",
            ]
        elif intent == "Commercial":
            templates = [
                f"Best {keyword_clean} Tools & Solutions in 2025",
                f"Top 10 {keyword_clean} Platforms Compared",
                f"{keyword_clean}: Complete Comparison & Review",
                f"How to Choose the Best {keyword_clean}",
                f"{keyword_clean} Review: Pros, Cons & Alternatives",
            ]
        elif intent == "Transactional":
            templates = [
                f"{keyword_clean}: Pricing, Plans & How to Get Started",
                f"Get Started with {keyword_clean} Today",
                f"{keyword_clean}: Features, Pricing & Free Trial",
                f"How to Buy {keyword_clean}: Complete Guide",
            ]
        else:  # Navigational
            templates = [
                f"{keyword_clean}: Official Guide & Resources",
                f"Getting Started with {keyword_clean}",
                f"{keyword_clean}: Quick Start Guide",
            ]
        
        # 返回第一个模板（可以根据需要随机选择）
        import hashlib
        hash_val = int(hashlib.md5(keyword.encode()).hexdigest(), 16)
        return templates[hash_val % len(templates)]
    
    def _cp_categorize_as_pillar(self, keyword: str, volume: int, difficulty: int) -> bool:
        """判断关键词是否适合作为支柱文章"""
        # 高搜索量 + 中等难度 = 支柱文章候选
        keyword_lower = keyword.lower()
        
        # 支柱文章特征：核心主题、高搜索量、竞争适中
        pillar_indicators = ['guide', 'complete', 'ultimate', 'what is', 'how to', 'best', 'top']
        has_pillar_indicator = any(ind in keyword_lower for ind in pillar_indicators)
        
        # 搜索量大于500，难度小于70
        volume_threshold = volume > 500 if isinstance(volume, int) else False
        difficulty_threshold = difficulty < 70 if isinstance(difficulty, int) else True
        
        return has_pillar_indicator or (volume_threshold and difficulty_threshold)

    # ==================== 公开工具方法 ====================

    def content_planning(
        self,
        domain: str,
        product_name: str,
        product_description: str,
        organic_keywords: str = "",
        related_keywords: str = "",
        question_keywords: str = "",
        keyword_research_file: str = "",
        article_count: int = 80,
        months: int = 6,
        start_date: str = ""
    ) -> str:
        """
        【内容规划工具】基于关键词研究结果生成Pillar-Based Marketing内容计划
        
        当用户说以下内容时调用此工具：
        - "内容规划"、"博客计划"、"文章规划"
        - "Pillar文章"、"支柱内容规划"
        - "规划6个月80篇文章"
        - "生成内容日历"
        - "基于关键词研究生成内容计划"
        
        :param domain: 网站域名（用于生成SEO友好的URL）
        :param product_name: 产品名称
        :param product_description: 产品/服务描述
        :param organic_keywords: 域名排名关键词（逗号分隔，格式：keyword:volume:difficulty）
        :param related_keywords: 相关关键词（逗号分隔）
        :param question_keywords: 问题型关键词（逗号分隔）
        :param keyword_research_file: 关键词研究Excel文件路径（可选，自动读取）
        :param article_count: 计划文章总数
        :param months: 规划周期（月）
        :param start_date: 开始日期（格式：YYYY-MM-DD，默认今天）
        :return: 包含Excel文件路径的结果
        """
        output_path = self._cp_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"content_plan_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        # 解析开始日期
        if start_date:
            try:
                plan_start_date = datetime.strptime(start_date, "%Y-%m-%d")
            except:
                plan_start_date = datetime.now()
        else:
            plan_start_date = datetime.now()
        
        # ===== 收集所有关键词 =====
        all_keywords = []
        
        # 1. 解析传入的关键词字符串
        def parse_keywords(kw_string: str, kw_type: str) -> List[dict]:
            keywords = []
            if not kw_string:
                return keywords
            for kw in kw_string.split(','):
                kw = kw.strip()
                if not kw:
                    continue
                # 支持格式：keyword:volume:difficulty 或 纯keyword
                parts = kw.split(':')
                keyword_data = {
                    'keyword': parts[0].strip(),
                    'volume': int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0,
                    'difficulty': int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 50,
                    'type': kw_type
                }
                keywords.append(keyword_data)
            return keywords
        
        all_keywords.extend(parse_keywords(organic_keywords, "Organic"))
        all_keywords.extend(parse_keywords(related_keywords, "Related"))
        all_keywords.extend(parse_keywords(question_keywords, "Question"))
        
        # 2. 如果提供了Excel文件，尝试读取
        if keyword_research_file and os.path.exists(keyword_research_file):
            try:
                kw_wb = load_workbook(keyword_research_file, read_only=True)
                for sheet_name in kw_wb.sheetnames:
                    if '说明' in sheet_name or '调试' in sheet_name:
                        continue
                    ws = kw_wb[sheet_name]
                    headers = [cell.value for cell in ws[1]]
                    
                    # 找到关键词和搜索量列
                    kw_col = None
                    vol_col = None
                    diff_col = None
                    
                    for idx, h in enumerate(headers):
                        if h and ('keyword' in str(h).lower() or 'ph' == str(h).lower()):
                            kw_col = idx
                        if h and ('volume' in str(h).lower() or 'nq' == str(h).lower()):
                            vol_col = idx
                        if h and ('difficult' in str(h).lower() or 'kd' == str(h).lower()):
                            diff_col = idx
                    
                    if kw_col is not None:
                        kw_type = "Organic" if "域名" in sheet_name else ("Question" if "问题" in sheet_name else "Related")
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[kw_col]:
                                all_keywords.append({
                                    'keyword': str(row[kw_col]),
                                    'volume': int(row[vol_col]) if vol_col and row[vol_col] else 0,
                                    'difficulty': int(row[diff_col]) if diff_col and row[diff_col] else 50,
                                    'type': kw_type
                                })
                kw_wb.close()
            except Exception as e:
                pass  # 文件读取失败，继续使用传入的关键词
        
        # 3. 如果没有关键词或关键词不足，生成更多关键词
        def generate_keywords_from_product(product: str, description: str, count: int) -> List[dict]:
            """根据产品名和描述生成足够数量的关键词"""
            generated = []
            product_lower = product.lower()
            
            # 核心模板（支柱文章候选）
            pillar_templates = [
                f"what is {product_lower}",
                f"complete guide to {product_lower}",
                f"best {product_lower} tools 2025",
                f"{product_lower} alternatives comparison",
                f"how to use {product_lower}",
                f"{product_lower} vs competitors",
                f"ultimate {product_lower} guide",
                f"{product_lower} for beginners",
            ]
            
            # 支撑文章模板
            cluster_templates = [
                # 教程类
                f"how to get started with {product_lower}",
                f"{product_lower} tutorial for beginners",
                f"step by step {product_lower} guide",
                f"{product_lower} tips and tricks",
                f"advanced {product_lower} techniques",
                f"{product_lower} best practices",
                f"common {product_lower} mistakes to avoid",
                f"{product_lower} workflow optimization",
                # 功能类
                f"{product_lower} features explained",
                f"{product_lower} pricing plans",
                f"{product_lower} integrations",
                f"{product_lower} api guide",
                f"{product_lower} automation tips",
                f"{product_lower} templates",
                f"{product_lower} plugins and extensions",
                # 对比类
                f"{product_lower} vs [competitor] comparison",
                f"is {product_lower} worth it",
                f"{product_lower} honest review 2025",
                f"{product_lower} pros and cons",
                f"why choose {product_lower}",
                f"{product_lower} case studies",
                f"{product_lower} success stories",
                # 问题类
                f"how does {product_lower} work",
                f"why use {product_lower}",
                f"when to use {product_lower}",
                f"who should use {product_lower}",
                f"what can {product_lower} do",
                f"is {product_lower} good for small business",
                f"can {product_lower} help with seo",
                f"does {product_lower} integrate with",
                # 场景类
                f"{product_lower} for marketing",
                f"{product_lower} for content creators",
                f"{product_lower} for agencies",
                f"{product_lower} for ecommerce",
                f"{product_lower} for startups",
                f"{product_lower} for enterprise",
                f"{product_lower} for freelancers",
                f"{product_lower} for bloggers",
                # 进阶类
                f"maximize roi with {product_lower}",
                f"{product_lower} analytics guide",
                f"measuring success with {product_lower}",
                f"{product_lower} reporting features",
                f"scaling with {product_lower}",
                f"{product_lower} team collaboration",
                f"{product_lower} security features",
                f"{product_lower} data privacy",
                # 行业相关
                f"{product_lower} industry trends",
                f"future of {product_lower}",
                f"{product_lower} updates 2025",
                f"{product_lower} roadmap",
                f"{product_lower} community",
                f"{product_lower} support resources",
                f"{product_lower} certification",
                f"{product_lower} training courses",
            ]
            
            # 从描述中提取关键词短语
            desc_words = description.lower().replace('，', ' ').replace(',', ' ').split()
            desc_phrases = []
            if len(desc_words) >= 2:
                for i in range(len(desc_words) - 1):
                    phrase = f"{desc_words[i]} {desc_words[i+1]}"
                    if len(phrase) > 5:
                        desc_phrases.append(phrase)
            
            # 基于描述生成更多关键词
            desc_templates = []
            for phrase in desc_phrases[:5]:
                desc_templates.extend([
                    f"best {phrase} tools",
                    f"how to improve {phrase}",
                    f"{phrase} strategies",
                    f"{phrase} tips 2025",
                    f"{phrase} guide for beginners",
                ])
            
            # 合并所有模板
            all_templates = pillar_templates + cluster_templates + desc_templates
            
            # 去重并生成关键词数据
            seen = set()
            for i, template in enumerate(all_templates):
                if len(generated) >= count:
                    break
                if template not in seen:
                    seen.add(template)
                    # 前8个作为高优先级（支柱候选）
                    is_pillar = i < 8
                    generated.append({
                        'keyword': template,
                        'volume': 500 if is_pillar else 100 + (i * 10) % 300,
                        'difficulty': 40 if is_pillar else 30 + (i * 5) % 40,
                        'type': 'Generated-Pillar' if is_pillar else 'Generated-Cluster'
                    })
            
            # 如果还不够，生成编号版本
            base_templates = [
                f"{product_lower} tip #{{n}}",
                f"{product_lower} strategy #{{n}}",
                f"{product_lower} use case #{{n}}",
                f"{product_lower} example #{{n}}",
                f"{product_lower} lesson #{{n}}",
            ]
            n = 1
            while len(generated) < count:
                for template in base_templates:
                    if len(generated) >= count:
                        break
                    kw = template.format(n=n)
                    if kw not in seen:
                        seen.add(kw)
                        generated.append({
                            'keyword': kw,
                            'volume': 50,
                            'difficulty': 25,
                            'type': 'Generated-Extended'
                        })
                n += 1
            
            return generated
        
        # 如果关键词不足，生成补充关键词
        if len(all_keywords) < article_count:
            needed = article_count - len(all_keywords)
            generated_keywords = generate_keywords_from_product(
                product_name, 
                product_description, 
                needed + 10  # 多生成一些以供筛选
            )
            all_keywords.extend(generated_keywords)
        
        # ===== 分类关键词为支柱和支撑 =====
        pillar_keywords = []
        cluster_keywords = []
        
        for kw_data in all_keywords:
            if self._cp_categorize_as_pillar(kw_data['keyword'], kw_data['volume'], kw_data['difficulty']):
                if len(pillar_keywords) < 8:  # 最多8个支柱
                    pillar_keywords.append(kw_data)
                else:
                    cluster_keywords.append(kw_data)
            else:
                cluster_keywords.append(kw_data)
        
        # 确保至少有5个支柱文章
        while len(pillar_keywords) < 5 and cluster_keywords:
            pillar_keywords.append(cluster_keywords.pop(0))
        
        # ===== 创建Excel =====
        wb = Workbook()
        
        # ===== Sheet 1: 完整内容计划 =====
        ws_main = wb.active
        ws_main.title = "内容发布计划"
        
        main_headers = [
            "编号", "文章类型", "发布日期", "发布周", "文章标题", 
            "目标关键词", "搜索意图", "SEO友好URL", "关键词来源",
            "搜索量", "难度", "预计字数", "状态"
        ]
        self._cp_style_excel_sheet(ws_main, main_headers, "2E75B6")
        
        # 计算发布节奏
        total_articles = min(article_count, len(pillar_keywords) + len(cluster_keywords))
        articles_per_week = max(1, total_articles // (months * 4))
        
        row_idx = 2
        article_num = 1
        current_date = plan_start_date
        week_num = 1
        
        # 先安排支柱文章（前几周优先发布）
        for i, kw_data in enumerate(pillar_keywords):
            keyword = kw_data['keyword']
            intent = self._cp_determine_search_intent(keyword)
            title = self._cp_generate_title_from_keyword(keyword, intent, product_name)
            url = self._cp_generate_seo_url(domain, keyword)
            
            # 支柱文章在前4周发布，每周1-2篇
            publish_date = plan_start_date + timedelta(days=i * 5)
            week_number = (i // 2) + 1
            
            ws_main.cell(row=row_idx, column=1, value=f"P{str(i+1).zfill(2)}")
            ws_main.cell(row=row_idx, column=2, value="支柱文章 Pillar")
            ws_main.cell(row=row_idx, column=3, value=publish_date.strftime("%Y-%m-%d"))
            ws_main.cell(row=row_idx, column=4, value=f"第{week_number}周")
            ws_main.cell(row=row_idx, column=5, value=title)
            ws_main.cell(row=row_idx, column=6, value=keyword)
            ws_main.cell(row=row_idx, column=7, value=intent)
            ws_main.cell(row=row_idx, column=8, value=url)
            ws_main.cell(row=row_idx, column=9, value=kw_data['type'])
            ws_main.cell(row=row_idx, column=10, value=kw_data['volume'])
            ws_main.cell(row=row_idx, column=11, value=kw_data['difficulty'])
            ws_main.cell(row=row_idx, column=12, value="2500-3500")
            ws_main.cell(row=row_idx, column=13, value="待写")
            row_idx += 1
            article_num += 1
        
        # 然后安排支撑文章
        pillar_end_date = plan_start_date + timedelta(days=len(pillar_keywords) * 5)
        cluster_start_week = (len(pillar_keywords) // 2) + 2
        
        for i, kw_data in enumerate(cluster_keywords):
            if article_num > article_count:
                break
                
            keyword = kw_data['keyword']
            intent = self._cp_determine_search_intent(keyword)
            title = self._cp_generate_title_from_keyword(keyword, intent, product_name)
            url = self._cp_generate_seo_url(domain, keyword)
            
            # 支撑文章均匀分布
            days_offset = (i // articles_per_week) * 7 + (i % articles_per_week) * 2
            publish_date = pillar_end_date + timedelta(days=days_offset)
            week_number = cluster_start_week + (i // articles_per_week)
            
            # 关联到支柱文章
            pillar_ref = f"P{str((i % len(pillar_keywords)) + 1).zfill(2)}"
            
            ws_main.cell(row=row_idx, column=1, value=f"C{str(i+1).zfill(3)}")
            ws_main.cell(row=row_idx, column=2, value=f"支撑文章 ({pillar_ref})")
            ws_main.cell(row=row_idx, column=3, value=publish_date.strftime("%Y-%m-%d"))
            ws_main.cell(row=row_idx, column=4, value=f"第{week_number}周")
            ws_main.cell(row=row_idx, column=5, value=title)
            ws_main.cell(row=row_idx, column=6, value=keyword)
            ws_main.cell(row=row_idx, column=7, value=intent)
            ws_main.cell(row=row_idx, column=8, value=url)
            ws_main.cell(row=row_idx, column=9, value=kw_data['type'])
            ws_main.cell(row=row_idx, column=10, value=kw_data['volume'])
            ws_main.cell(row=row_idx, column=11, value=kw_data['difficulty'])
            ws_main.cell(row=row_idx, column=12, value="1200-1800")
            ws_main.cell(row=row_idx, column=13, value="待写")
            row_idx += 1
            article_num += 1
        
        actual_articles = row_idx - 2
        
        # 调整列宽
        ws_main.column_dimensions['A'].width = 8
        ws_main.column_dimensions['B'].width = 18
        ws_main.column_dimensions['C'].width = 12
        ws_main.column_dimensions['D'].width = 10
        ws_main.column_dimensions['E'].width = 50
        ws_main.column_dimensions['F'].width = 35
        ws_main.column_dimensions['G'].width = 15
        ws_main.column_dimensions['H'].width = 55
        ws_main.column_dimensions['I'].width = 12
        ws_main.column_dimensions['J'].width = 10
        ws_main.column_dimensions['K'].width = 8
        ws_main.column_dimensions['L'].width = 12
        ws_main.column_dimensions['M'].width = 10
        
        # ===== Sheet 2: 发布日历视图 =====
        ws_calendar = wb.create_sheet("发布日历")
        
        cal_headers = ["周次", "日期范围", "支柱文章", "支撑文章", "总计", "累计发布"]
        self._cp_style_excel_sheet(ws_calendar, cal_headers, "ED7D31")
        
        # 按周统计
        week_stats = {}
        for row in ws_main.iter_rows(min_row=2, max_row=actual_articles + 1, values_only=True):
            week = row[3]  # 发布周
            article_type = row[1]  # 文章类型
            if week not in week_stats:
                week_stats[week] = {'pillar': 0, 'cluster': 0}
            if '支柱' in str(article_type):
                week_stats[week]['pillar'] += 1
            else:
                week_stats[week]['cluster'] += 1
        
        cumulative = 0
        cal_row = 2
        for week in sorted(week_stats.keys(), key=lambda x: int(x.replace('第', '').replace('周', ''))):
            week_num = int(week.replace('第', '').replace('周', ''))
            week_start = plan_start_date + timedelta(days=(week_num - 1) * 7)
            week_end = week_start + timedelta(days=6)
            
            pillar_count = week_stats[week]['pillar']
            cluster_count = week_stats[week]['cluster']
            total = pillar_count + cluster_count
            cumulative += total
            
            ws_calendar.cell(row=cal_row, column=1, value=week)
            ws_calendar.cell(row=cal_row, column=2, value=f"{week_start.strftime('%m/%d')} - {week_end.strftime('%m/%d')}")
            ws_calendar.cell(row=cal_row, column=3, value=pillar_count)
            ws_calendar.cell(row=cal_row, column=4, value=cluster_count)
            ws_calendar.cell(row=cal_row, column=5, value=total)
            ws_calendar.cell(row=cal_row, column=6, value=cumulative)
            cal_row += 1
        
        # ===== Sheet 3: 支柱-支撑映射 =====
        ws_mapping = wb.create_sheet("支柱支撑映射")
        
        map_headers = ["支柱文章编号", "支柱文章标题", "关联支撑文章数", "支撑文章编号列表"]
        self._cp_style_excel_sheet(ws_mapping, map_headers, "7030A0")
        
        # 统计每个支柱的支撑文章
        pillar_clusters = {f"P{str(i+1).zfill(2)}": [] for i in range(len(pillar_keywords))}
        for row in ws_main.iter_rows(min_row=2, max_row=actual_articles + 1, values_only=True):
            if '支撑' in str(row[1]):
                # 提取关联的支柱编号
                import re
                match = re.search(r'P\d+', str(row[1]))
                if match:
                    pillar_id = match.group()
                    if pillar_id in pillar_clusters:
                        pillar_clusters[pillar_id].append(row[0])
        
        map_row = 2
        for i, kw_data in enumerate(pillar_keywords):
            pillar_id = f"P{str(i+1).zfill(2)}"
            pillar_title = self._cp_generate_title_from_keyword(kw_data['keyword'], 
                           self._cp_determine_search_intent(kw_data['keyword']), product_name)
            clusters = pillar_clusters.get(pillar_id, [])
            
            ws_mapping.cell(row=map_row, column=1, value=pillar_id)
            ws_mapping.cell(row=map_row, column=2, value=pillar_title)
            ws_mapping.cell(row=map_row, column=3, value=len(clusters))
            ws_mapping.cell(row=map_row, column=4, value=", ".join(clusters[:10]) + ("..." if len(clusters) > 10 else ""))
            map_row += 1
        
        ws_mapping.column_dimensions['B'].width = 50
        ws_mapping.column_dimensions['D'].width = 40
        
        # ===== Sheet 4: 关键词来源统计 =====
        ws_stats = wb.create_sheet("关键词统计")
        
        stats_headers = ["关键词来源", "数量", "占比"]
        self._cp_style_excel_sheet(ws_stats, stats_headers, "70AD47")
        
        source_counts = {}
        for kw in all_keywords:
            source = kw['type']
            source_counts[source] = source_counts.get(source, 0) + 1
        
        stats_row = 2
        total_kw = len(all_keywords)
        for source, count in source_counts.items():
            ws_stats.cell(row=stats_row, column=1, value=source)
            ws_stats.cell(row=stats_row, column=2, value=count)
            ws_stats.cell(row=stats_row, column=3, value=f"{count/total_kw*100:.1f}%")
            stats_row += 1
        
        wb.save(filepath)
        
        return f"""
📊 **Pillar-Based 内容规划完成**

🌐 网站: {domain}
📦 产品: {product_name}
📅 开始日期: {plan_start_date.strftime("%Y-%m-%d")}

═══════════════════════════════════════
📈 **规划统计**
═══════════════════════════════════════
📅 规划周期: {months} 个月
📄 计划文章总数: {actual_articles} 篇
  - 🏛️ 支柱文章 (Pillar): {len(pillar_keywords)} 篇
  - 📝 支撑文章 (Cluster): {actual_articles - len(pillar_keywords)} 篇
🔑 关键词来源:
{chr(10).join([f'  - {k}: {v}个' for k, v in source_counts.items()])}

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. **内容发布计划** - 完整文章列表
   - 具体发布日期
   - SEO友好URL
   - 搜索意图分析
   - 关键词数据
2. **发布日历** - 按周统计的发布节奏
3. **支柱支撑映射** - Pillar与Cluster的关联关系
4. **关键词统计** - 关键词来源分布

💡 **使用说明**:
- URL 已根据关键词自动生成，格式为: {domain}/blog/[seo-slug]
- 搜索意图已自动分析（Informational/Commercial/Transactional）
- 支柱文章优先在前4周发布
- 支撑文章按周均匀分布
"""

    def geo_optimization_plan(
        self,
        domain: str,
        product_name: str,
        product_description: str,
        prompt_count: int = 40
    ) -> str:
        """
        【GEO优化计划工具】制定GEO（AI回答优化）运营计划，生成AI提示词和监测方法
        
        当用户说以下内容时调用此工具：
        - "GEO优化"、"GEO计划"、"AI优化"
        - "AI回答优化"、"AI搜索优化"
        - "生成40个AI提示词"
        - "制定GEO运营计划"
        
        :param domain: 网站域名
        :param product_name: 产品名称
        :param product_description: 产品/服务描述
        :param prompt_count: 生成的AI提示词数量
        :return: 包含Excel文件路径的结果
        """
        output_path = self._cp_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"geo_plan_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        wb = Workbook()
        
        # ===== Sheet 1: AI提示词库 =====
        ws1 = wb.active
        ws1.title = "AI提示词库"
        
        prompt_headers = ["编号", "提示词类型", "AI提示词", "目标AI平台", "预期回答包含", "优先级", "监测频率", "当前状态"]
        self._cp_style_excel_sheet(ws1, prompt_headers, "2E75B6")
        
        # 提示词类型和模板
        prompt_types = [
            ("产品推荐", [
                f"What is the best {product_name.lower()} tool?",
                f"Recommend a {product_name.lower()} solution for my business",
                f"Top {product_name.lower()} platforms in 2025",
                f"Which {product_name.lower()} should I use?",
                f"Best {product_name.lower()} for beginners",
                f"Best {product_name.lower()} for enterprise",
                f"Most affordable {product_name.lower()} tools",
                f"Best free {product_name.lower()} alternatives",
            ]),
            ("产品对比", [
                f"{product_name} vs [Competitor] comparison",
                f"Is {product_name} better than [Alternative]?",
                f"Compare {product_name} with other solutions",
                f"{product_name} alternatives comparison",
                f"Pros and cons of {product_name}",
            ]),
            ("问题解决", [
                f"How to solve [problem] with {product_name}?",
                f"Can {product_name} help with [use case]?",
                f"How does {product_name} work?",
                f"Getting started with {product_name}",
                f"How to set up {product_name}?",
            ]),
            ("使用场景", [
                f"Best {product_name.lower()} for [industry]",
                f"How to use {product_name} for [purpose]?",
                f"{product_name} use cases and examples",
                f"Who should use {product_name}?",
            ]),
            ("评价评测", [
                f"Is {product_name} worth it?",
                f"{product_name} honest review",
                f"What do users say about {product_name}?",
                f"{product_name} pricing review",
            ]),
        ]
        
        platforms = ["ChatGPT", "Perplexity", "Claude", "Gemini", "Google AI Overview"]
        priorities = ["高", "高", "中", "中", "低"]
        frequencies = ["每周", "每周", "每两周", "每月", "每月"]
        
        row_idx = 2
        prompt_num = 1
        
        for type_name, prompts in prompt_types:
            for prompt in prompts:
                if prompt_num > prompt_count:
                    break
                ws1.cell(row=row_idx, column=1, value=f"GEO-{str(prompt_num).zfill(3)}")
                ws1.cell(row=row_idx, column=2, value=type_name)
                ws1.cell(row=row_idx, column=3, value=prompt)
                ws1.cell(row=row_idx, column=4, value=platforms[(prompt_num - 1) % len(platforms)])
                ws1.cell(row=row_idx, column=5, value=f"提及 {product_name}")
                ws1.cell(row=row_idx, column=6, value=priorities[(prompt_num - 1) % len(priorities)])
                ws1.cell(row=row_idx, column=7, value=frequencies[(prompt_num - 1) % len(frequencies)])
                ws1.cell(row=row_idx, column=8, value="待监测")
                row_idx += 1
                prompt_num += 1
        
        # ===== Sheet 2: 监测记录 =====
        ws2 = wb.create_sheet("监测记录")
        
        monitor_headers = ["日期", "提示词编号", "AI平台", "是否被提及", "排名位置", "回答摘要", "优化建议"]
        self._cp_style_excel_sheet(ws2, monitor_headers, "70AD47")
        
        # ===== Sheet 3: 优化策略 =====
        ws3 = wb.create_sheet("优化策略")
        
        strategy_headers = ["策略类型", "具体措施", "预期效果", "执行周期", "负责人", "状态"]
        self._cp_style_excel_sheet(ws3, strategy_headers, "ED7D31")
        
        strategies = [
            ("内容优化", "在文章开头40字内直接回答核心问题", "提高AI引用概率", "立即执行", "", "待执行"),
            ("内容优化", "添加TL;DR摘要和FAQ结构", "便于AI提取信息", "立即执行", "", "待执行"),
            ("内容优化", "增加原创数据、统计和案例", "增强内容可信度", "持续进行", "", "待执行"),
            ("技术优化", "添加Schema.org结构化数据", "帮助AI理解内容结构", "1周内", "", "待执行"),
            ("技术优化", "优化页面加载速度", "提高爬取效率", "2周内", "", "待执行"),
            ("外部建设", "获取权威网站引用和反向链接", "提升网站权威性", "持续进行", "", "待执行"),
            ("外部建设", "在问答平台建立品牌存在", "增加品牌曝光", "持续进行", "", "待执行"),
            ("监测分析", "每周监测目标提示词的AI回答", "跟踪优化效果", "每周", "", "待执行"),
        ]
        
        for row_idx, data in enumerate(strategies, 2):
            for col_idx, value in enumerate(data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        
        return f"""
📊 **GEO优化计划完成**

🌐 网站: {domain}
📦 产品: {product_name}
📝 描述: {product_description}

═══════════════════════════════════════
📈 **计划统计**
═══════════════════════════════════════
🤖 AI提示词数量: {min(prompt_num - 1, prompt_count)} 个
📊 覆盖场景:
  - 产品推荐类提示词
  - 产品对比类提示词
  - 问题解决类提示词
  - 使用场景类提示词
  - 评价评测类提示词
🎯 目标平台: ChatGPT, Perplexity, Claude, Gemini, Google AI Overview

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. AI提示词库 - {min(prompt_num - 1, prompt_count)}个目标提示词及监测计划
2. 监测记录 - 用于记录每次监测结果
3. 优化策略 - 8条GEO优化具体措施

💡 **GEO优化核心要点**:
- 内容开头直接回答问题（前40字）
- 使用清晰的结构（TL;DR, FAQ, 列表）
- 提供原创数据和统计信息
- 添加结构化数据标记
- 持续监测和优化
"""

    def project_task_list(
        self,
        project_name: str,
        domain: str,
        include_seo: bool = True,
        include_geo: bool = True
    ) -> str:
        """
        【项目任务清单工具】生成完整的SEO+GEO项目运营任务清单
        
        当用户说以下内容时调用此工具：
        - "任务清单"、"项目管理"、"任务列表"
        - "甘特图"、"项目计划"
        - "生成工作清单"
        - "SEO/GEO项目任务"
        
        :param project_name: 项目名称
        :param domain: 网站域名
        :param include_seo: 是否包含SEO任务
        :param include_geo: 是否包含GEO任务
        :return: 包含Excel文件路径的结果
        """
        output_path = self._cp_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"project_tasks_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        wb = Workbook()
        
        # ===== Sheet 1: 任务总览 =====
        ws1 = wb.active
        ws1.title = "任务总览"
        
        task_headers = ["任务ID", "阶段", "任务名称", "描述", "优先级", "预计工时", "开始日期", "截止日期", "负责人", "状态", "完成度"]
        self._cp_style_excel_sheet(ws1, task_headers, "2E75B6")
        
        start_date = datetime.now()
        tasks = []
        task_id = 1
        
        # 准备阶段任务
        prep_tasks = [
            ("准备阶段", "项目启动会议", "与客户确认项目目标和范围", "高", "2h"),
            ("准备阶段", "网站访问权限获取", "获取Google Analytics、Search Console等权限", "高", "1h"),
            ("准备阶段", "竞品分析", "分析3-5个主要竞争对手", "高", "4h"),
            ("准备阶段", "现状审计", "网站SEO现状全面检查", "高", "4h"),
        ]
        
        for stage, name, desc, priority, hours in prep_tasks:
            days_offset = (task_id - 1) * 2
            tasks.append((
                f"T{str(task_id).zfill(3)}",
                stage, name, desc, priority, hours,
                (start_date + timedelta(days=days_offset)).strftime("%Y-%m-%d"),
                (start_date + timedelta(days=days_offset + 2)).strftime("%Y-%m-%d"),
                "", "待开始", "0%"
            ))
            task_id += 1
        
        if include_seo:
            seo_tasks = [
                ("SEO执行", "关键词研究", "挖掘目标关键词并分类", "高", "8h"),
                ("SEO执行", "页面关键词映射", "分配关键词到对应页面", "高", "4h"),
                ("SEO执行", "内容规划制定", "制定6个月内容发布计划", "高", "4h"),
                ("SEO执行", "技术SEO优化", "修复技术问题，优化网站结构", "中", "8h"),
                ("SEO执行", "内容创作-支柱文章", "撰写5篇支柱文章", "高", "20h"),
                ("SEO执行", "内容创作-支撑文章", "持续产出支撑文章", "中", "持续"),
                ("SEO执行", "内部链接优化", "建立内容之间的链接关系", "中", "4h"),
                ("SEO执行", "外部链接建设", "获取高质量外链", "中", "持续"),
            ]
            
            for stage, name, desc, priority, hours in seo_tasks:
                days_offset = 7 + (task_id - 5) * 3
                tasks.append((
                    f"T{str(task_id).zfill(3)}",
                    stage, name, desc, priority, hours,
                    (start_date + timedelta(days=days_offset)).strftime("%Y-%m-%d"),
                    (start_date + timedelta(days=days_offset + 7)).strftime("%Y-%m-%d"),
                    "", "待开始", "0%"
                ))
                task_id += 1
        
        if include_geo:
            geo_tasks = [
                ("GEO执行", "GEO提示词规划", "制定40+个目标AI提示词", "高", "4h"),
                ("GEO执行", "内容GEO优化", "优化现有内容以适应AI引用", "高", "8h"),
                ("GEO执行", "结构化数据添加", "添加Schema.org标记", "中", "4h"),
                ("GEO执行", "AI平台监测设置", "建立监测流程和工具", "中", "2h"),
                ("GEO执行", "GEO效果追踪", "每周监测AI回答中的品牌露出", "中", "持续"),
            ]
            
            for stage, name, desc, priority, hours in geo_tasks:
                days_offset = 14 + (task_id - len(tasks) - 1) * 3
                tasks.append((
                    f"T{str(task_id).zfill(3)}",
                    stage, name, desc, priority, hours,
                    (start_date + timedelta(days=days_offset)).strftime("%Y-%m-%d"),
                    (start_date + timedelta(days=days_offset + 7)).strftime("%Y-%m-%d"),
                    "", "待开始", "0%"
                ))
                task_id += 1
        
        # 交付阶段任务
        delivery_tasks = [
            ("交付阶段", "月度报告", "整理月度SEO/GEO数据报告", "中", "4h"),
            ("交付阶段", "效果复盘", "分析优化效果，调整策略", "中", "2h"),
            ("交付阶段", "项目总结", "输出项目总结文档", "低", "2h"),
        ]
        
        for stage, name, desc, priority, hours in delivery_tasks:
            days_offset = 30 + (task_id - len(tasks) - 1) * 7
            tasks.append((
                f"T{str(task_id).zfill(3)}",
                stage, name, desc, priority, hours,
                (start_date + timedelta(days=days_offset)).strftime("%Y-%m-%d"),
                (start_date + timedelta(days=days_offset + 7)).strftime("%Y-%m-%d"),
                "", "待开始", "0%"
            ))
            task_id += 1
        
        for row_idx, data in enumerate(tasks, 2):
            for col_idx, value in enumerate(data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        
        # ===== Sheet 2: 里程碑 =====
        ws2 = wb.create_sheet("里程碑")
        
        milestone_headers = ["里程碑", "描述", "目标日期", "关联任务", "状态"]
        self._cp_style_excel_sheet(ws2, milestone_headers, "70AD47")
        
        milestones = [
            ("M1: 项目启动", "完成启动会议和权限获取", (start_date + timedelta(days=3)).strftime("%Y-%m-%d"), "T001-T002", "待完成"),
            ("M2: 研究完成", "完成关键词研究和竞品分析", (start_date + timedelta(days=14)).strftime("%Y-%m-%d"), "T003-T006", "待完成"),
            ("M3: 规划完成", "完成内容规划和GEO计划", (start_date + timedelta(days=21)).strftime("%Y-%m-%d"), "T007-T010", "待完成"),
            ("M4: 首批内容上线", "发布首批5篇支柱文章", (start_date + timedelta(days=45)).strftime("%Y-%m-%d"), "T009", "待完成"),
            ("M5: 项目复盘", "完成首月效果分析", (start_date + timedelta(days=60)).strftime("%Y-%m-%d"), "T016-T017", "待完成"),
        ]
        
        for row_idx, data in enumerate(milestones, 2):
            for col_idx, value in enumerate(data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        
        seo_count = len([t for t in tasks if "SEO" in t[1]]) if include_seo else 0
        geo_count = len([t for t in tasks if "GEO" in t[1]]) if include_geo else 0
        
        return f"""
📊 **项目任务清单完成**

📁 项目名称: {project_name}
🌐 网站: {domain}

═══════════════════════════════════════
📈 **任务统计**
═══════════════════════════════════════
📋 总任务数: {len(tasks)} 项
  - 准备阶段: 4 项
  - SEO执行: {seo_count} 项
  - GEO执行: {geo_count} 项
  - 交付阶段: 3 项
🎯 里程碑: 5 个

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. 任务总览 - 完整任务列表（含时间、负责人、状态）
2. 里程碑 - 关键节点和检查点

💡 **使用建议**:
- 每周更新任务状态和完成度
- 关注里程碑的按时完成
- 根据实际情况调整优先级
"""

