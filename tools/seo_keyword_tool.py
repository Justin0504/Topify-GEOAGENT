"""
title: SEO 关键词研究工具
description: 【关键词研究】分析网站并扩展SEO关键词 | 【页面关键词映射】生成页面与关键词的对应关系表
author: GEO Agent
version: 1.0.0
required_open_webui_version: 0.6.0
requirements: openpyxl, requests, urllib3
"""

import os
import requests
import urllib3
from typing import Dict, Any, List, Optional
from datetime import datetime
from pydantic import BaseModel, Field

# 禁用 SSL 警告（某些 Docker 环境可能需要）
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class Tools:
    """
    SEO 关键词研究工具 - 使用 Semrush API 进行关键词分析
    
    ═══════════════════════════════════════════════════════════════
    🎯 功能匹配指南（中文触发词）
    ═══════════════════════════════════════════════════════════════
    
    🔍 keyword_research - 关键词研究
       触发词: "关键词研究", "扩展关键词", "SEO关键词", "分析关键词", 
              "找关键词", "挖掘关键词", "关键词分析"
       示例: "帮我对 topify.ai 做关键词研究"
       输出: Excel 文件（包含搜索量、难度、CPC等数据）
    
    📄 page_keyword_mapping - 页面关键词映射
       触发词: "页面映射", "关键词映射", "页面优化", "关键词分配",
              "哪个页面优化哪些词", "页面关键词对应"
       示例: "生成页面-关键词映射表"
       输出: Excel 文件（页面URL与对应关键词）
    
    ═══════════════════════════════════════════════════════════════
    """

    class Valves(BaseModel):
        OUTPUT_PATH: str = Field(
            default="/app/backend/data/output",
            description="文件保存路径（Docker环境）"
        )
        SEMRUSH_API_KEY: str = Field(
            default="",
            description="【必填】Semrush API Key"
        )
        DEFAULT_DATABASE: str = Field(
            default="us",
            description="默认数据库/地区 (us=美国, cn=中国, uk=英国, jp=日本)"
        )
        DEFAULT_LIMIT: int = Field(
            default=100,
            description="默认返回关键词数量上限"
        )

    def __init__(self):
        self.valves = self.Valves()
        self.api_base_url = "https://api.semrush.com/"

    # ==================== 私有方法 ====================

    def _kw_ensure_output_dir(self) -> str:
        """确保输出目录存在"""
        output_path = self.valves.OUTPUT_PATH
        if not os.path.exists(output_path):
            os.makedirs(output_path, exist_ok=True)
        return output_path

    def _kw_make_api_request(self, params: dict) -> dict:
        """
        发送 Semrush API 请求
        """
        api_key = self.valves.SEMRUSH_API_KEY.strip()
        
        if not api_key:
            return {
                "success": False,
                "error": "❌ 未配置 Semrush API Key，请在工具设置(Valves)中配置",
                "data": []
            }
        
        params["key"] = api_key
        
        try:
            # verify=False 用于处理某些 Docker 环境的 SSL 证书问题
            response = requests.get(self.api_base_url, params=params, timeout=30, verify=False)
            raw_text = response.text.strip()
            
            if raw_text.startswith("ERROR"):
                error_code = raw_text.split("::")[0] if "::" in raw_text else raw_text
                error_messages = {
                    "ERROR 50": "API Key 无效或已过期",
                    "ERROR 40": "超出 API 调用限制",
                    "ERROR 120": "无效的数据库/地区代码",
                    "ERROR 130": "请求的数据库中没有此数据"
                }
                parts = error_code.split(" ")
                key = f"{parts[0]} {parts[1]}" if len(parts) > 1 else error_code
                friendly_error = error_messages.get(key, raw_text)
                return {"success": False, "error": f"API 错误: {friendly_error}", "data": []}
            
            lines = raw_text.split("\n")
            if not lines or not lines[0]:
                return {"success": True, "data": [], "columns": [], "count": 0}
            
            columns = lines[0].split(";")
            data = []
            for line in lines[1:]:
                if line.strip():
                    values = line.split(";")
                    row = dict(zip(columns, values))
                    data.append(row)
            
            return {
                "success": True,
                "data": data,
                "columns": columns,
                "count": len(data),
                "raw_text": raw_text
            }
            
        except requests.exceptions.Timeout:
            return {"success": False, "error": "请求超时，请稍后重试", "data": []}
        except requests.exceptions.RequestException as e:
            return {"success": False, "error": f"网络错误: {str(e)}", "data": []}
        except Exception as e:
            return {"success": False, "error": f"解析错误: {str(e)}", "data": []}

    def _kw_get_domain_keywords(self, domain: str, database: str = None, limit: int = None) -> dict:
        """获取域名排名关键词"""
        params = {
            "type": "domain_organic",
            "domain": domain,
            "database": database or self.valves.DEFAULT_DATABASE,
            "display_limit": limit or self.valves.DEFAULT_LIMIT,
            "export_columns": "Ph,Po,Nq,Cp,Co,Kd,Tr,Tc,Nr,Td,Ur"
        }
        return self._kw_make_api_request(params)

    def _kw_get_related_keywords(self, keyword: str, database: str = None, limit: int = None) -> dict:
        """获取相关关键词"""
        params = {
            "type": "phrase_related",
            "phrase": keyword,
            "database": database or self.valves.DEFAULT_DATABASE,
            "display_limit": limit or self.valves.DEFAULT_LIMIT,
            "export_columns": "Ph,Nq,Cp,Co,Kd,Nr,Td"
        }
        return self._kw_make_api_request(params)

    def _kw_get_question_keywords(self, keyword: str, database: str = None, limit: int = None) -> dict:
        """获取问题型关键词"""
        params = {
            "type": "phrase_questions",
            "phrase": keyword,
            "database": database or self.valves.DEFAULT_DATABASE,
            "display_limit": limit or self.valves.DEFAULT_LIMIT,
            "export_columns": "Ph,Nq,Cp,Co,Kd"
        }
        return self._kw_make_api_request(params)

    def _kw_style_excel_sheet(self, ws, headers: List[str]):
        """为Excel表格添加样式"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # ==================== 公开工具方法 ====================

    def keyword_research(
        self,
        domain: str,
        product_description: str = "",
        target_market: str = "us",
        limit: int = 100
    ) -> str:
        """
        【关键词研究工具】分析网站并扩展所有产品相关的SEO关键词
        
        当用户说以下内容时调用此工具：
        - "关键词研究"、"扩展关键词"、"SEO关键词分析"
        - "帮我分析 xxx 网站的关键词"
        - "找出所有产品相关的关键词"
        - "做关键词挖掘"
        
        :param domain: 要分析的网站域名，例如 "topify.ai" 或 "example.com"
        :param product_description: 产品/服务描述，帮助筛选相关关键词
        :param target_market: 目标市场地区代码 (us=美国, cn=中国, uk=英国, jp=日本)
        :param limit: 每类关键词的数量上限
        :return: 包含Excel文件路径和关键词统计的结果
        """
        output_path = self._kw_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"keyword_research_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        wb = Workbook()
        results_summary = []
        
        # 1. 获取域名现有排名关键词
        domain_result = self._kw_get_domain_keywords(domain, target_market, limit)
        ws1 = wb.active
        ws1.title = "域名排名关键词"
        
        if domain_result["success"] and domain_result.get("data"):
            columns = domain_result.get("columns", [])
            self._kw_style_excel_sheet(ws1, columns)
            for row_idx, row_data in enumerate(domain_result["data"], 2):
                for col_idx, col in enumerate(columns, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
            results_summary.append(f"✅ 域名排名关键词: {domain_result['count']} 个")
        else:
            ws1.cell(row=1, column=1, value=f"获取失败: {domain_result.get('error', '未知错误')}")
            results_summary.append(f"❌ 域名排名关键词: {domain_result.get('error', '获取失败')}")
        
        # 2. 确定用于扩展的核心关键词
        # 优先使用产品描述，其次使用域名
        if product_description:
            # 从产品描述中提取关键词（取前3-4个词）
            core_keyword = ' '.join(product_description.replace('，', ' ').replace(',', ' ').split()[:4])
        else:
            core_keyword = domain.split('.')[0]
        
        # 获取相关关键词
        related_result = self._kw_get_related_keywords(core_keyword, target_market, limit)
        ws2 = wb.create_sheet("相关关键词")
        
        if related_result["success"] and related_result.get("data"):
            columns = related_result.get("columns", [])
            self._kw_style_excel_sheet(ws2, columns)
            for row_idx, row_data in enumerate(related_result["data"], 2):
                for col_idx, col in enumerate(columns, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
            results_summary.append(f"✅ 相关关键词: {related_result['count']} 个 (基于: {core_keyword})")
        else:
            # 如果产品描述的关键词没找到，尝试用通用的 SEO 相关词
            fallback_keyword = "seo optimization"
            related_result = self._kw_get_related_keywords(fallback_keyword, target_market, limit)
            if related_result["success"] and related_result.get("data"):
                columns = related_result.get("columns", [])
                self._kw_style_excel_sheet(ws2, columns)
                for row_idx, row_data in enumerate(related_result["data"], 2):
                    for col_idx, col in enumerate(columns, 1):
                        ws2.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
                results_summary.append(f"✅ 相关关键词: {related_result['count']} 个 (基于: {fallback_keyword})")
            else:
                ws2.cell(row=1, column=1, value=f"获取失败: {related_result.get('error', '未知错误')}")
                results_summary.append(f"❌ 相关关键词: {related_result.get('error', '获取失败')}")
        
        # 3. 获取问题型关键词
        question_result = self._kw_get_question_keywords(core_keyword, target_market, limit)
        ws3 = wb.create_sheet("问题型关键词")
        
        if question_result["success"] and question_result.get("data"):
            columns = question_result.get("columns", [])
            self._kw_style_excel_sheet(ws3, columns)
            for row_idx, row_data in enumerate(question_result["data"], 2):
                for col_idx, col in enumerate(columns, 1):
                    ws3.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
            results_summary.append(f"✅ 问题型关键词: {question_result['count']} 个")
        else:
            # 如果没找到，尝试用通用的 AI SEO 相关问题
            fallback_keyword = "ai seo"
            question_result = self._kw_get_question_keywords(fallback_keyword, target_market, limit)
            if question_result["success"] and question_result.get("data"):
                columns = question_result.get("columns", [])
                self._kw_style_excel_sheet(ws3, columns)
                for row_idx, row_data in enumerate(question_result["data"], 2):
                    for col_idx, col in enumerate(columns, 1):
                        ws3.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
                results_summary.append(f"✅ 问题型关键词: {question_result['count']} 个 (基于: {fallback_keyword})")
            else:
                ws3.cell(row=1, column=1, value=f"获取失败: {question_result.get('error', '未知错误')}")
                results_summary.append(f"❌ 问题型关键词: {question_result.get('error', '获取失败')}")
        
        # 4. 添加说明页
        ws_info = wb.create_sheet("字段说明")
        info_data = [
            ["字段", "说明"],
            ["Ph", "关键词 (Keyword)"],
            ["Po", "排名位置 (Position)"],
            ["Nq", "月搜索量 (Search Volume)"],
            ["Cp", "CPC 点击成本 (Cost Per Click)"],
            ["Co", "竞争度 (Competition)"],
            ["Kd", "关键词难度 (Keyword Difficulty)"],
            ["Tr", "流量 (Traffic)"],
            ["Tc", "流量成本 (Traffic Cost)"],
            ["Nr", "搜索结果数 (Number of Results)"],
            ["Td", "趋势 (Trend)"],
            ["Ur", "排名URL (URL)"]
        ]
        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_info.cell(row=row_idx, column=col_idx, value=value)
        
        # 保存文件
        wb.save(filepath)
        
        total_keywords = sum([
            domain_result.get("count", 0),
            related_result.get("count", 0),
            question_result.get("count", 0)
        ])
        
        return f"""
📊 **关键词研究完成**

🌐 分析域名: {domain}
🎯 目标市场: {target_market}
📝 产品描述: {product_description or '未提供'}

═══════════════════════════════════════
📈 **数据统计**
═══════════════════════════════════════
{chr(10).join(results_summary)}

📁 **总计**: {total_keywords} 个关键词

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. 域名排名关键词 - 网站当前排名的关键词
2. 相关关键词 - 与核心词相关的拓展词
3. 问题型关键词 - 用户常问的问题形式关键词
4. 字段说明 - 数据字段解释
"""

    def _kw_get_domain_pages(self, domain: str, database: str = None, limit: int = None) -> dict:
        """获取域名的页面列表（通过 domain_organic_organic 获取所有排名页面）"""
        params = {
            "type": "domain_organic_organic",
            "domain": domain,
            "database": database or self.valves.DEFAULT_DATABASE,
            "display_limit": limit or self.valves.DEFAULT_LIMIT,
            "export_columns": "Dn,Ur"  # 只获取域名和URL
        }
        return self._kw_make_api_request(params)

    def _kw_get_url_keywords(self, url: str, database: str = None, limit: int = None) -> dict:
        """获取指定URL的排名关键词"""
        params = {
            "type": "url_organic",
            "url": url,
            "database": database or self.valves.DEFAULT_DATABASE,
            "display_limit": limit or 50,
            "export_columns": "Ph,Po,Nq,Cp,Co,Kd,Tr"
        }
        return self._kw_make_api_request(params)

    def page_keyword_mapping(
        self,
        domain: str,
        target_market: str = "us",
        limit: int = 200
    ) -> str:
        """
        【页面关键词映射工具】生成页面-关键词映射表，明确每个页面应该优化哪些关键词
        
        当用户说以下内容时调用此工具：
        - "页面映射"、"关键词映射"、"页面关键词对应"
        - "每个页面优化哪些关键词"
        - "生成页面-关键词映射表"
        - "关键词分配到页面"
        
        :param domain: 要分析的网站域名
        :param target_market: 目标市场地区代码
        :param limit: 返回关键词数量上限
        :return: 包含Excel文件路径的结果
        """
        output_path = self._kw_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"page_keyword_mapping_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        wb = Workbook()
        debug_info = []
        
        # 方法1: 获取域名关键词（包含URL信息）
        result = self._kw_get_domain_keywords(domain, target_market, limit)
        debug_info.append(f"domain_organic API: success={result.get('success')}, count={result.get('count', 0)}")
        
        # 检查返回的列名
        if result.get("columns"):
            debug_info.append(f"返回字段: {', '.join(result['columns'])}")
        
        # Sheet 1: 原始API数据
        ws_raw = wb.active
        ws_raw.title = "原始API数据"
        
        url_keywords = {}
        page_count = 0
        keyword_count = 0
        
        if result["success"] and result.get("data"):
            # 写入原始数据
            columns = result.get("columns", [])
            if columns:
                self._kw_style_excel_sheet(ws_raw, columns)
                for row_idx, row_data in enumerate(result["data"], 2):
                    for col_idx, col in enumerate(columns, 1):
                        ws_raw.cell(row=row_idx, column=col_idx, value=row_data.get(col, ""))
            
            # 按URL分组
            for item in result["data"]:
                # Semrush 的 URL 字段可能是 "Ur" 或 "Url"
                url = item.get("Ur") or item.get("Url") or item.get("url") or ""
                if not url:
                    url = f"https://{domain}/"  # 如果没有URL，默认为首页
                
                if url not in url_keywords:
                    url_keywords[url] = []
                
                url_keywords[url].append({
                    "keyword": item.get("Ph") or item.get("Keyword") or "",
                    "position": item.get("Po") or item.get("Position") or "",
                    "volume": item.get("Nq") or item.get("Search Volume") or "0",
                    "difficulty": item.get("Kd") or item.get("Keyword Difficulty") or "",
                    "traffic": item.get("Tr") or item.get("Traffic") or "0",
                    "cpc": item.get("Cp") or item.get("CPC") or ""
                })
            
            page_count = len(url_keywords)
            keyword_count = result["count"]
            debug_info.append(f"分组后页面数: {page_count}")
            
        else:
            error_msg = result.get('error', '未知错误')
            debug_info.append(f"API 错误: {error_msg}")
            
            # 如果 domain_organic 失败，尝试创建模板
            ws_raw.cell(row=1, column=1, value="API说明")
            ws_raw.cell(row=2, column=1, value=f"域名 {domain} 在 Semrush {target_market} 数据库中暂无排名数据")
            ws_raw.cell(row=3, column=1, value="可能原因: 1) 新网站 2) 小众市场 3) 数据库选择不对")
            ws_raw.cell(row=4, column=1, value="建议: 1) 尝试其他地区代码 2) 先进行关键词研究再手动映射")
        
        # Sheet 2: 页面关键词映射（格式化后的数据）
        ws_mapping = wb.create_sheet("页面关键词映射")
        headers = ["页面URL", "关键词", "当前排名", "月搜索量", "难度", "预估流量", "CPC"]
        self._kw_style_excel_sheet(ws_mapping, headers)
        
        if url_keywords:
            row_idx = 2
            for url, keywords in sorted(url_keywords.items(), key=lambda x: len(x[1]), reverse=True):
                for kw in keywords:
                    ws_mapping.cell(row=row_idx, column=1, value=url)
                    ws_mapping.cell(row=row_idx, column=2, value=kw["keyword"])
                    ws_mapping.cell(row=row_idx, column=3, value=kw["position"])
                    ws_mapping.cell(row=row_idx, column=4, value=kw["volume"])
                    ws_mapping.cell(row=row_idx, column=5, value=kw["difficulty"])
                    ws_mapping.cell(row=row_idx, column=6, value=kw["traffic"])
                    ws_mapping.cell(row=row_idx, column=7, value=kw["cpc"])
                    row_idx += 1
        else:
            # 创建空模板供用户手动填写
            ws_mapping.cell(row=2, column=1, value=f"https://{domain}/")
            ws_mapping.cell(row=2, column=2, value="(请手动添加目标关键词)")
            ws_mapping.cell(row=3, column=1, value=f"https://{domain}/about")
            ws_mapping.cell(row=3, column=2, value="(请手动添加目标关键词)")
            ws_mapping.cell(row=4, column=1, value=f"https://{domain}/product")
            ws_mapping.cell(row=4, column=2, value="(请手动添加目标关键词)")
        
        # Sheet 3: 页面汇总
        ws_summary = wb.create_sheet("页面汇总")
        summary_headers = ["页面URL", "关键词数量", "总预估流量", "平均排名"]
        self._kw_style_excel_sheet(ws_summary, summary_headers)
        
        if url_keywords:
            row_idx = 2
            for url, keywords in sorted(url_keywords.items(), key=lambda x: len(x[1]), reverse=True):
                # 计算总流量
                total_traffic = 0
                valid_positions = []
                for kw in keywords:
                    traffic_str = str(kw["traffic"]).replace(",", "").strip()
                    if traffic_str.isdigit():
                        total_traffic += int(traffic_str)
                    pos_str = str(kw["position"]).strip()
                    if pos_str.isdigit():
                        valid_positions.append(int(pos_str))
                
                avg_position = round(sum(valid_positions) / len(valid_positions), 1) if valid_positions else "-"
                
                ws_summary.cell(row=row_idx, column=1, value=url)
                ws_summary.cell(row=row_idx, column=2, value=len(keywords))
                ws_summary.cell(row=row_idx, column=3, value=total_traffic)
                ws_summary.cell(row=row_idx, column=4, value=avg_position)
                row_idx += 1
        
        # Sheet 4: 调试信息
        ws_debug = wb.create_sheet("调试信息")
        ws_debug.cell(row=1, column=1, value="调试日志")
        for idx, info in enumerate(debug_info, 2):
            ws_debug.cell(row=idx, column=1, value=info)
        
        wb.save(filepath)
        
        # 构建结果消息
        if keyword_count > 0:
            status_msg = "✅ 数据获取成功"
        else:
            status_msg = "⚠️ 未获取到排名数据（已生成空模板）"
        
        return f"""
📊 **页面关键词映射完成**

🌐 分析域名: {domain}
🎯 目标市场: {target_market}
📌 状态: {status_msg}

═══════════════════════════════════════
📈 **统计数据**
═══════════════════════════════════════
📄 页面数量: {page_count}
🔑 关键词总数: {keyword_count}

═══════════════════════════════════════
🔍 **调试信息**
═══════════════════════════════════════
{chr(10).join(debug_info)}

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. 原始API数据 - Semrush 返回的原始数据
2. 页面关键词映射 - 格式化后的映射表
3. 页面汇总 - 各页面的统计信息
4. 调试信息 - API 调用日志

💡 **如果数据为空**:
- 尝试更换 target_market 参数（如 cn, uk, de）
- 新网站可能需要等待 Semrush 收录
- 可以先用 keyword_research 找到目标关键词，再手动分配到页面
"""

