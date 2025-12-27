"""
title: 技术SEO审计工具
description: 【技术SEO检查】扫描网站技术SEO问题，区分需修复和需优化的问题，并给出具体建议
author: GEO Agent
version: 1.0.0
required_open_webui_version: 0.6.0
requirements: openpyxl, requests, beautifulsoup4
"""

import os
import requests
from typing import Dict, Any, List, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from urllib.parse import urljoin, urlparse

try:
    from bs4 import BeautifulSoup
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class Tools:
    """
    技术SEO审计工具 - 扫描网站技术问题并生成审计报告
    
    ═══════════════════════════════════════════════════════════════
    🎯 功能匹配指南（中文触发词）
    ═══════════════════════════════════════════════════════════════
    
    🔧 technical_seo_audit - 技术SEO审计
       触发词: "技术SEO", "SEO检查", "网站审计", "SEO审计",
              "技术问题", "SEO问题扫描", "网站诊断"
       示例: "对 example.com 进行技术SEO检查"
       输出: Excel 文件（问题列表和修复建议）
    
    ═══════════════════════════════════════════════════════════════
    """

    class Valves(BaseModel):
        OUTPUT_PATH: str = Field(
            default="/app/backend/data/output",
            description="文件保存路径（Docker环境）"
        )
        REQUEST_TIMEOUT: int = Field(
            default=10,
            description="HTTP请求超时时间（秒）"
        )

    def __init__(self):
        self.valves = self.Valves()

    def _ts_ensure_output_dir(self) -> str:
        """确保输出目录存在"""
        output_path = self.valves.OUTPUT_PATH
        if not os.path.exists(output_path):
            os.makedirs(output_path, exist_ok=True)
        return output_path

    def _ts_style_excel_sheet(self, ws, headers: List[str], header_color: str = "4472C4"):
        """为Excel表格添加样式"""
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.freeze_panes = 'A2'

    def _fetch_page(self, url: str) -> dict:
        """获取页面内容"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (compatible; SEOBot/1.0; +http://example.com/bot)'
            }
            response = requests.get(url, headers=headers, timeout=self.valves.REQUEST_TIMEOUT, allow_redirects=True)
            return {
                "success": True,
                "status_code": response.status_code,
                "content": response.text,
                "headers": dict(response.headers),
                "url": response.url,
                "elapsed": response.elapsed.total_seconds()
            }
        except requests.exceptions.Timeout:
            return {"success": False, "error": "请求超时"}
        except requests.exceptions.RequestException as e:
            return {"success": False, "error": str(e)}

    def _analyze_page(self, url: str, html_content: str, headers: dict) -> List[dict]:
        """分析页面SEO问题"""
        issues = []
        
        if not BS4_AVAILABLE:
            issues.append({
                "category": "系统",
                "severity": "警告",
                "issue": "BeautifulSoup未安装",
                "description": "无法进行深度HTML分析",
                "recommendation": "安装 beautifulsoup4: pip install beautifulsoup4",
                "impact": "中"
            })
            return issues
        
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 1. 检查 Title 标签
        title = soup.find('title')
        if not title:
            issues.append({
                "category": "Meta标签",
                "severity": "严重",
                "issue": "缺少Title标签",
                "description": "页面没有<title>标签",
                "recommendation": "添加唯一、描述性的title标签（50-60字符）",
                "impact": "高"
            })
        elif title:
            title_text = title.get_text().strip()
            if len(title_text) < 30:
                issues.append({
                    "category": "Meta标签",
                    "severity": "警告",
                    "issue": "Title过短",
                    "description": f"Title仅{len(title_text)}字符: '{title_text}'",
                    "recommendation": "Title建议50-60字符，包含主要关键词",
                    "impact": "中"
                })
            elif len(title_text) > 60:
                issues.append({
                    "category": "Meta标签",
                    "severity": "建议",
                    "issue": "Title过长",
                    "description": f"Title有{len(title_text)}字符，可能被截断",
                    "recommendation": "Title建议50-60字符",
                    "impact": "低"
                })
        
        # 2. 检查 Meta Description
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        if not meta_desc:
            issues.append({
                "category": "Meta标签",
                "severity": "严重",
                "issue": "缺少Meta Description",
                "description": "页面没有meta description标签",
                "recommendation": "添加描述性的meta description（150-160字符）",
                "impact": "高"
            })
        elif meta_desc:
            desc_content = meta_desc.get('content', '').strip()
            if len(desc_content) < 70:
                issues.append({
                    "category": "Meta标签",
                    "severity": "警告",
                    "issue": "Meta Description过短",
                    "description": f"描述仅{len(desc_content)}字符",
                    "recommendation": "建议150-160字符，包含关键词和号召性用语",
                    "impact": "中"
                })
            elif len(desc_content) > 160:
                issues.append({
                    "category": "Meta标签",
                    "severity": "建议",
                    "issue": "Meta Description过长",
                    "description": f"描述有{len(desc_content)}字符，可能被截断",
                    "recommendation": "建议150-160字符",
                    "impact": "低"
                })
        
        # 3. 检查 H1 标签
        h1_tags = soup.find_all('h1')
        if len(h1_tags) == 0:
            issues.append({
                "category": "标题结构",
                "severity": "严重",
                "issue": "缺少H1标签",
                "description": "页面没有H1标签",
                "recommendation": "每个页面应有且仅有一个H1标签",
                "impact": "高"
            })
        elif len(h1_tags) > 1:
            issues.append({
                "category": "标题结构",
                "severity": "警告",
                "issue": "多个H1标签",
                "description": f"页面有{len(h1_tags)}个H1标签",
                "recommendation": "每个页面应仅有一个H1标签",
                "impact": "中"
            })
        
        # 4. 检查图片 Alt 属性
        images = soup.find_all('img')
        images_without_alt = [img for img in images if not img.get('alt')]
        if images_without_alt:
            issues.append({
                "category": "图片优化",
                "severity": "警告",
                "issue": "图片缺少Alt属性",
                "description": f"{len(images_without_alt)}/{len(images)}张图片缺少alt属性",
                "recommendation": "为所有图片添加描述性的alt文本",
                "impact": "中"
            })
        
        # 5. 检查 Canonical 标签
        canonical = soup.find('link', attrs={'rel': 'canonical'})
        if not canonical:
            issues.append({
                "category": "技术SEO",
                "severity": "警告",
                "issue": "缺少Canonical标签",
                "description": "页面没有canonical标签",
                "recommendation": "添加canonical标签避免重复内容问题",
                "impact": "中"
            })
        
        # 6. 检查 Robots Meta
        robots_meta = soup.find('meta', attrs={'name': 'robots'})
        if robots_meta:
            content = robots_meta.get('content', '').lower()
            if 'noindex' in content:
                issues.append({
                    "category": "索引控制",
                    "severity": "严重",
                    "issue": "页面设置为noindex",
                    "description": "页面被设置为不索引",
                    "recommendation": "如非故意，移除noindex指令",
                    "impact": "高"
                })
        
        # 7. 检查移动端viewport
        viewport = soup.find('meta', attrs={'name': 'viewport'})
        if not viewport:
            issues.append({
                "category": "移动端优化",
                "severity": "严重",
                "issue": "缺少Viewport meta标签",
                "description": "页面没有设置viewport",
                "recommendation": "添加: <meta name='viewport' content='width=device-width, initial-scale=1'>",
                "impact": "高"
            })
        
        # 8. 检查 HTTPS
        if not url.startswith('https://'):
            issues.append({
                "category": "安全性",
                "severity": "严重",
                "issue": "未使用HTTPS",
                "description": "网站未使用HTTPS加密",
                "recommendation": "配置SSL证书，启用HTTPS",
                "impact": "高"
            })
        
        # 9. 检查结构化数据
        schema_scripts = soup.find_all('script', attrs={'type': 'application/ld+json'})
        if not schema_scripts:
            issues.append({
                "category": "结构化数据",
                "severity": "建议",
                "issue": "缺少结构化数据",
                "description": "页面没有Schema.org结构化数据",
                "recommendation": "添加适合内容类型的JSON-LD结构化数据",
                "impact": "中"
            })
        
        # 10. 检查内部链接
        internal_links = soup.find_all('a', href=True)
        broken_links = []
        for link in internal_links[:10]:  # 只检查前10个链接
            href = link.get('href', '')
            if href.startswith('/') or url in href:
                # 这是内部链接，可以进一步检查
                pass
        
        # 11. 检查页面加载相关
        if headers:
            # 检查压缩
            if 'gzip' not in headers.get('Content-Encoding', '').lower():
                issues.append({
                    "category": "性能优化",
                    "severity": "建议",
                    "issue": "未启用Gzip压缩",
                    "description": "页面响应未使用gzip压缩",
                    "recommendation": "在服务器配置中启用gzip压缩",
                    "impact": "中"
                })
            
            # 检查缓存
            cache_control = headers.get('Cache-Control', '')
            if not cache_control:
                issues.append({
                    "category": "性能优化",
                    "severity": "建议",
                    "issue": "未设置缓存策略",
                    "description": "响应头没有Cache-Control",
                    "recommendation": "设置适当的缓存策略",
                    "impact": "低"
                })
        
        return issues

    def technical_seo_audit(
        self,
        domain: str,
        pages_to_check: str = ""
    ) -> str:
        """
        【技术SEO审计工具】扫描网站技术SEO问题，生成详细审计报告
        
        当用户说以下内容时调用此工具：
        - "技术SEO检查"、"SEO审计"、"网站诊断"
        - "扫描SEO问题"、"技术问题检查"
        - "网站SEO健康检查"
        
        :param domain: 要检查的网站域名（如 example.com）
        :param pages_to_check: 要额外检查的页面路径（逗号分隔，如 "/about,/contact"）
        :return: 包含Excel审计报告路径的结果
        """
        output_path = self._ts_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"technical_seo_audit_{domain.replace('.', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        # 构建URL列表
        base_url = f"https://{domain}" if not domain.startswith('http') else domain
        urls_to_check = [base_url]
        
        if pages_to_check:
            for page in pages_to_check.split(','):
                page = page.strip()
                if page:
                    urls_to_check.append(urljoin(base_url, page))
        
        wb = Workbook()
        all_issues = []
        page_results = []
        
        # 检查每个页面
        for url in urls_to_check:
            result = self._fetch_page(url)
            
            if result["success"]:
                issues = self._analyze_page(url, result["content"], result.get("headers", {}))
                for issue in issues:
                    issue["page"] = url
                all_issues.extend(issues)
                
                page_results.append({
                    "url": url,
                    "status": result["status_code"],
                    "load_time": f"{result['elapsed']:.2f}s",
                    "issues_count": len(issues)
                })
            else:
                all_issues.append({
                    "category": "可访问性",
                    "severity": "严重",
                    "issue": "页面无法访问",
                    "description": result.get("error", "未知错误"),
                    "recommendation": "检查URL是否正确，服务器是否正常运行",
                    "impact": "高",
                    "page": url
                })
                page_results.append({
                    "url": url,
                    "status": "错误",
                    "load_time": "-",
                    "issues_count": 1
                })
        
        # ===== Sheet 1: 问题汇总 =====
        ws1 = wb.active
        ws1.title = "问题汇总"
        
        issue_headers = ["问题ID", "页面", "类别", "严重程度", "问题", "描述", "修复建议", "影响程度"]
        self._ts_style_excel_sheet(ws1, issue_headers, "C00000")
        
        # 按严重程度排序
        severity_order = {"严重": 0, "警告": 1, "建议": 2}
        sorted_issues = sorted(all_issues, key=lambda x: severity_order.get(x.get("severity", "建议"), 3))
        
        for row_idx, issue in enumerate(sorted_issues, 2):
            ws1.cell(row=row_idx, column=1, value=f"SEO-{str(row_idx-1).zfill(3)}")
            ws1.cell(row=row_idx, column=2, value=issue.get("page", ""))
            ws1.cell(row=row_idx, column=3, value=issue.get("category", ""))
            ws1.cell(row=row_idx, column=4, value=issue.get("severity", ""))
            ws1.cell(row=row_idx, column=5, value=issue.get("issue", ""))
            ws1.cell(row=row_idx, column=6, value=issue.get("description", ""))
            ws1.cell(row=row_idx, column=7, value=issue.get("recommendation", ""))
            ws1.cell(row=row_idx, column=8, value=issue.get("impact", ""))
            
            # 根据严重程度着色
            severity = issue.get("severity", "")
            if severity == "严重":
                ws1.cell(row=row_idx, column=4).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws1.cell(row=row_idx, column=4).font = Font(color="FFFFFF", bold=True)
            elif severity == "警告":
                ws1.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            elif severity == "建议":
                ws1.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 调整列宽
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 25
        ws1.column_dimensions['F'].width = 40
        ws1.column_dimensions['G'].width = 50
        ws1.column_dimensions['H'].width = 12
        
        # ===== Sheet 2: 页面检查结果 =====
        ws2 = wb.create_sheet("页面检查结果")
        
        page_headers = ["页面URL", "HTTP状态", "加载时间", "问题数量"]
        self._ts_style_excel_sheet(ws2, page_headers, "2E75B6")
        
        for row_idx, page in enumerate(page_results, 2):
            ws2.cell(row=row_idx, column=1, value=page["url"])
            ws2.cell(row=row_idx, column=2, value=page["status"])
            ws2.cell(row=row_idx, column=3, value=page["load_time"])
            ws2.cell(row=row_idx, column=4, value=page["issues_count"])
        
        ws2.column_dimensions['A'].width = 50
        
        # ===== Sheet 3: 修复优先级 =====
        ws3 = wb.create_sheet("修复优先级")
        
        priority_headers = ["优先级", "类别", "问题数量", "建议操作"]
        self._ts_style_excel_sheet(ws3, priority_headers, "70AD47")
        
        # 统计各类问题
        category_counts = {}
        for issue in all_issues:
            cat = issue.get("category", "其他")
            if cat not in category_counts:
                category_counts[cat] = {"严重": 0, "警告": 0, "建议": 0}
            category_counts[cat][issue.get("severity", "建议")] += 1
        
        row_idx = 2
        
        # 先处理严重问题
        for cat, counts in category_counts.items():
            if counts["严重"] > 0:
                ws3.cell(row=row_idx, column=1, value="立即修复")
                ws3.cell(row=row_idx, column=2, value=cat)
                ws3.cell(row=row_idx, column=3, value=counts["严重"])
                ws3.cell(row=row_idx, column=4, value="这些问题严重影响SEO，需要立即处理")
                ws3.cell(row=row_idx, column=1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws3.cell(row=row_idx, column=1).font = Font(color="FFFFFF", bold=True)
                row_idx += 1
        
        # 再处理警告问题
        for cat, counts in category_counts.items():
            if counts["警告"] > 0:
                ws3.cell(row=row_idx, column=1, value="尽快优化")
                ws3.cell(row=row_idx, column=2, value=cat)
                ws3.cell(row=row_idx, column=3, value=counts["警告"])
                ws3.cell(row=row_idx, column=4, value="这些问题影响SEO效果，建议在2周内处理")
                ws3.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                row_idx += 1
        
        # 最后处理建议
        for cat, counts in category_counts.items():
            if counts["建议"] > 0:
                ws3.cell(row=row_idx, column=1, value="持续改进")
                ws3.cell(row=row_idx, column=2, value=cat)
                ws3.cell(row=row_idx, column=3, value=counts["建议"])
                ws3.cell(row=row_idx, column=4, value="这些是优化建议，可以在日常维护中处理")
                ws3.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                row_idx += 1
        
        wb.save(filepath)
        
        # 统计
        critical_count = len([i for i in all_issues if i.get("severity") == "严重"])
        warning_count = len([i for i in all_issues if i.get("severity") == "警告"])
        suggestion_count = len([i for i in all_issues if i.get("severity") == "建议"])
        
        return f"""
📊 **技术SEO审计完成**

🌐 网站: {domain}
📄 检查页面数: {len(urls_to_check)}

═══════════════════════════════════════
📈 **问题统计**
═══════════════════════════════════════
🔴 严重问题: {critical_count} 个（需立即修复）
🟠 警告问题: {warning_count} 个（建议2周内优化）
🟡 改进建议: {suggestion_count} 个（持续优化）

📋 **总计**: {len(all_issues)} 个问题

═══════════════════════════════════════
💾 **文件已保存**
═══════════════════════════════════════
路径: {filepath}

📋 包含以下工作表:
1. 问题汇总 - 所有问题详情及修复建议
2. 页面检查结果 - 各页面状态和加载时间
3. 修复优先级 - 按优先级分类的行动建议

💡 **优化建议**:
1. 优先处理"严重"问题（红色标记）
2. 重点关注 Meta标签、标题结构、移动端优化
3. 定期进行技术SEO检查（建议每月一次）
"""

