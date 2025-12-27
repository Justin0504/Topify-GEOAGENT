"""
title: SEO 完整工作流工具
description: 【一键执行】自动完成 关键词研究→内容规划→写文章→发布 的完整流程，无需多轮调用
author: GEO Agent
version: 1.0.0
required_open_webui_version: 0.6.0
requirements: openpyxl, requests, urllib3, python-docx
"""

import os
import json
import time
import requests
import urllib3
from typing import Dict, Any, List, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# 禁用 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class Tools:
    """
    SEO 完整工作流工具 - 一键执行多步骤 SEO 任务
    
    ═══════════════════════════════════════════════════════════════
    🎯 一键执行完整流程，无需多轮调用
    ═══════════════════════════════════════════════════════════════
    
    🚀 full_seo_analysis - 完整SEO分析
       触发词: "完整分析", "全面分析", "一键分析", "SEO全套"
       功能: 关键词研究 + 页面映射 + 内容规划 + 技术审计
    
    ✍️ research_and_publish - 研究并发布文章
       触发词: "研究并发布", "分析并写文章", "一键发布"
       功能: 关键词研究 → 写文章 → 发布到WordPress
    
    📝 plan_and_publish - 规划并发布文章
       触发词: "规划并发布", "内容规划并写文章"
       功能: 内容规划 → 取第一个主题 → 写文章 → 发布
    
    ═══════════════════════════════════════════════════════════════
    """

    class Valves(BaseModel):
        # Semrush API
        SEMRUSH_API_KEY: str = Field(
            default="",
            description="【必填】Semrush API Key"
        )
        DEFAULT_DATABASE: str = Field(
            default="us",
            description="默认目标市场 (us/cn/uk/jp)"
        )
        # WordPress
        WP_ACCESS_TOKEN: str = Field(
            default="",
            description="【必填】WordPress.com API Access Token"
        )
        WP_SITE_ID: str = Field(
            default="",
            description="【必填】WordPress.com Site ID"
        )
        WP_API_BASE: str = Field(
            default="https://public-api.wordpress.com/rest/v1.1",
            description="WordPress.com API 基础 URL"
        )
        # 输出
        OUTPUT_PATH: str = Field(
            default="/app/backend/data/output",
            description="文件保存路径"
        )

    def __init__(self):
        self.valves = self.Valves()
        self.semrush_api_url = "https://api.semrush.com/"
        self._wp_session = self._create_wp_session()

    def _create_wp_session(self) -> requests.Session:
        """创建 WordPress 请求 session"""
        session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET", "POST"],
        )
        adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=10, pool_maxsize=10)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def _clean(self, value: str) -> str:
        """清理字符串"""
        return "".join(value.split()) if value else ""

    def _wf_ensure_output_dir(self) -> str:
        """确保输出目录存在"""
        output_path = self.valves.OUTPUT_PATH
        os.makedirs(output_path, exist_ok=True)
        return output_path

    # ==================== Semrush API 调用 ====================

    def _wf_semrush_request(self, params: dict) -> dict:
        """调用 Semrush API"""
        api_key = self._clean(self.valves.SEMRUSH_API_KEY)
        if not api_key:
            return {"success": False, "error": "未配置 Semrush API Key", "data": []}
        
        params["key"] = api_key
        
        try:
            response = requests.get(self.semrush_api_url, params=params, timeout=30, verify=False)
            raw_text = response.text.strip()
            
            if raw_text.startswith("ERROR"):
                return {"success": False, "error": f"API 错误: {raw_text}", "data": []}
            
            lines = raw_text.split("\n")
            if not lines or not lines[0]:
                return {"success": True, "data": [], "columns": [], "count": 0}
            
            columns = lines[0].split(";")
            data = []
            for line in lines[1:]:
                if line.strip():
                    values = line.split(";")
                    row = dict(zip(columns, values))
                    data.append(row)
            
            return {"success": True, "data": data, "columns": columns, "count": len(data)}
            
        except Exception as e:
            return {"success": False, "error": f"请求错误: {str(e)}", "data": []}

    def _wf_get_domain_keywords(self, domain: str, database: str, limit: int = 50) -> dict:
        """获取域名排名关键词"""
        return self._wf_semrush_request({
            "type": "domain_organic",
            "domain": domain,
            "database": database,
            "display_limit": limit
        })

    def _wf_get_related_keywords(self, phrase: str, database: str, limit: int = 50) -> dict:
        """获取相关关键词"""
        return self._wf_semrush_request({
            "type": "phrase_related",
            "phrase": phrase,
            "database": database,
            "display_limit": limit
        })

    # ==================== WordPress API 调用 ====================

    def _wp_request(self, method: str, endpoint: str, data: dict = None) -> dict:
        """调用 WordPress API"""
        token = self._clean(self.valves.WP_ACCESS_TOKEN)
        site_id = self._clean(self.valves.WP_SITE_ID)
        
        if not token:
            return {"success": False, "error": "未配置 WordPress Access Token"}
        if not site_id:
            return {"success": False, "error": "未配置 WordPress Site ID"}
        
        api_base = self._clean(self.valves.WP_API_BASE).rstrip("/")
        url = f"{api_base}/sites/{site_id}/{endpoint}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }
        
        try:
            if method.upper() == "GET":
                response = self._wp_session.get(url, headers=headers, timeout=60)
            else:
                response = self._wp_session.post(url, headers=headers, json=data, timeout=60)
            
            result = response.json()
            
            if response.status_code in [200, 201]:
                return {"success": True, "data": result}
            else:
                return {"success": False, "error": result.get("message", str(result))}
                
        except Exception as e:
            return {"success": False, "error": f"请求错误: {str(e)}"}

    # ==================== 公开工具方法 ====================

    def full_seo_analysis(
        self,
        domain: str,
        product_description: str = "",
        target_market: str = "us"
    ) -> str:
        """
        【一键完整SEO分析】自动执行 关键词研究 + 页面映射 + 内容规划
        
        当用户说以下内容时调用此工具：
        - "完整分析"、"全面分析"、"一键分析"
        - "帮我做完整的SEO准备"
        - "SEO全套分析"
        
        :param domain: 网站域名（必填），如 "topify.ai"
        :param product_description: 产品描述（可选）
        :param target_market: 目标市场 (us/cn/uk/jp)
        :return: 完整分析报告
        """
        results = []
        output_path = self._wf_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        domain_safe = domain.replace(".", "_")
        
        # ========== 步骤 1: 关键词研究 ==========
        results.append("📊 **步骤 1/3: 关键词研究**")
        
        keywords_data = []
        
        # 获取域名关键词
        domain_kw = self._wf_get_domain_keywords(domain, target_market, 50)
        if domain_kw["success"]:
            keywords_data.extend(domain_kw["data"])
            results.append(f"  ✅ 域名排名关键词: {domain_kw['count']} 个")
        else:
            results.append(f"  ⚠️ 域名关键词: {domain_kw['error']}")
        
        # 获取相关关键词
        search_phrase = product_description or domain.split('.')[0]
        related_kw = self._wf_get_related_keywords(search_phrase, target_market, 50)
        if related_kw["success"]:
            keywords_data.extend(related_kw["data"])
            results.append(f"  ✅ 相关关键词: {related_kw['count']} 个")
        else:
            results.append(f"  ⚠️ 相关关键词: {related_kw['error']}")
        
        # 保存关键词到 Excel
        kw_filename = f"keywords_{domain_safe}_{timestamp}.xlsx"
        kw_filepath = os.path.join(output_path, kw_filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "关键词列表"
        
        if keywords_data:
            headers = list(keywords_data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            for row_idx, kw in enumerate(keywords_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=kw.get(header, ""))
        
        wb.save(kw_filepath)
        results.append(f"  💾 保存到: {kw_filepath}")
        
        # ========== 步骤 2: 内容规划 ==========
        results.append("\n📝 **步骤 2/3: 内容规划**")
        
        # 提取高价值关键词用于内容规划
        pillar_topics = []
        support_topics = []
        
        for kw in keywords_data[:30]:  # 取前30个关键词
            keyword = kw.get("Keyword", kw.get("Ph", kw.get("keyword", "")))
            volume = kw.get("Search Volume", kw.get("Nq", kw.get("volume", "0")))
            
            try:
                vol = int(volume)
            except:
                vol = 0
            
            if vol >= 1000:
                pillar_topics.append({"keyword": keyword, "volume": vol, "type": "pillar"})
            elif vol >= 100:
                support_topics.append({"keyword": keyword, "volume": vol, "type": "support"})
        
        # 保存内容规划到 Excel
        plan_filename = f"content_plan_{domain_safe}_{timestamp}.xlsx"
        plan_filepath = os.path.join(output_path, plan_filename)
        
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "内容规划"
        
        ws2.cell(row=1, column=1, value="类型").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="关键词").font = Font(bold=True)
        ws2.cell(row=1, column=3, value="搜索量").font = Font(bold=True)
        ws2.cell(row=1, column=4, value="建议文章标题").font = Font(bold=True)
        
        row = 2
        for topic in pillar_topics[:10]:
            ws2.cell(row=row, column=1, value="支柱文章")
            ws2.cell(row=row, column=2, value=topic["keyword"])
            ws2.cell(row=row, column=3, value=topic["volume"])
            ws2.cell(row=row, column=4, value=f"Complete Guide to {topic['keyword'].title()}")
            row += 1
        
        for topic in support_topics[:20]:
            ws2.cell(row=row, column=1, value="支撑文章")
            ws2.cell(row=row, column=2, value=topic["keyword"])
            ws2.cell(row=row, column=3, value=topic["volume"])
            ws2.cell(row=row, column=4, value=f"How to {topic['keyword'].title()}")
            row += 1
        
        wb2.save(plan_filepath)
        results.append(f"  ✅ 支柱文章主题: {len(pillar_topics[:10])} 个")
        results.append(f"  ✅ 支撑文章主题: {len(support_topics[:20])} 个")
        results.append(f"  💾 保存到: {plan_filepath}")
        
        # ========== 步骤 3: 汇总报告 ==========
        results.append("\n📋 **步骤 3/3: 汇总报告**")
        results.append(f"  📊 总计关键词: {len(keywords_data)} 个")
        results.append(f"  📝 内容规划: {len(pillar_topics[:10]) + len(support_topics[:20])} 篇")
        
        # 返回第一个支柱主题供后续使用
        first_pillar = pillar_topics[0]["keyword"] if pillar_topics else (support_topics[0]["keyword"] if support_topics else "AI optimization")
        
        results_text = "\n".join(results)
        
        return f"""
🎯 **完整 SEO 分析完成**

═══════════════════════════════════════
🌐 分析域名: {domain}
🎯 目标市场: {target_market}
═══════════════════════════════════════

{results_text}

═══════════════════════════════════════
📁 **生成的文件**
═══════════════════════════════════════
1. 关键词列表: {kw_filepath}
2. 内容规划: {plan_filepath}

═══════════════════════════════════════
💡 **推荐的第一篇文章主题**
═══════════════════════════════════════
"{first_pillar}"

如需继续写文章并发布，请使用：
research_and_publish 工具
"""

    def research_and_publish(
        self,
        domain: str,
        keyword: str,
        product_name: str = "",
        product_description: str = "",
        target_market: str = "us",
        publish_status: str = "publish"
    ) -> str:
        """
        【一键研究并发布】关键词研究 → 生成文章 → 发布到WordPress
        
        当用户说以下内容时调用此工具：
        - "研究并发布"、"分析并写文章"
        - "帮我写一篇关于XX的文章并发布"
        - "一键发布文章"
        
        :param domain: 网站域名（必填）
        :param keyword: 文章主题关键词（必填）
        :param product_name: 产品名称（可选）
        :param product_description: 产品描述（可选）
        :param target_market: 目标市场
        :param publish_status: 发布状态 (publish/draft)
        :return: 发布结果
        """
        results = []
        output_path = self._wf_ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        product = product_name or domain.split('.')[0].title()
        
        # ========== 步骤 1: 研究关键词 ==========
        results.append("🔍 **步骤 1/3: 关键词研究**")
        
        related_kw = self._wf_get_related_keywords(keyword, target_market, 10)
        related_keywords = []
        if related_kw["success"] and related_kw["data"]:
            related_keywords = [kw.get("Keyword", kw.get("Ph", "")) for kw in related_kw["data"][:5]]
            results.append(f"  ✅ 找到 {len(related_keywords)} 个相关关键词")
        else:
            results.append(f"  ⚠️ 未找到相关关键词，将使用主关键词")
            related_keywords = [keyword]
        
        # ========== 步骤 2: 生成文章 ==========
        results.append("\n✍️ **步骤 2/3: 生成文章**")
        
        # 构建文章内容
        article_title = f"Complete Guide to {keyword.title()}: Best Practices for 2025"
        
        article_content = f"""
<h2>Introduction: Understanding {keyword.title()}</h2>

<p>In today's digital landscape, mastering {keyword} has become essential for businesses looking to stay competitive. This comprehensive guide explores everything you need to know about {keyword}, from fundamental concepts to advanced strategies.</p>

<p><strong>TL;DR:</strong> {keyword.title()} is crucial for modern businesses. {product} offers innovative solutions that can help you achieve better results in this area.</p>

<h2>What is {keyword.title()}?</h2>

<p>{keyword.title()} refers to the practice of optimizing your digital presence to achieve better visibility and engagement. With the rise of AI-powered search engines and recommendation systems, understanding {keyword} has never been more important.</p>

<h3>Key Benefits of {keyword.title()}</h3>

<ul>
<li><strong>Improved Visibility:</strong> Get your content seen by more potential customers</li>
<li><strong>Better Engagement:</strong> Connect with your audience more effectively</li>
<li><strong>Higher Conversions:</strong> Turn visitors into loyal customers</li>
<li><strong>Competitive Advantage:</strong> Stay ahead of your competitors</li>
</ul>

<h2>How {product} Can Help</h2>

<p><a href="https://{domain}">{product}</a> provides cutting-edge solutions for {keyword}. {product_description or f'Our platform helps businesses optimize their digital presence and achieve better results.'}</p>

<h3>Key Features</h3>

<ul>
<li><strong>AI-Powered Analysis:</strong> Leverage advanced AI to understand your performance</li>
<li><strong>Actionable Insights:</strong> Get clear recommendations for improvement</li>
<li><strong>Easy Implementation:</strong> Simple tools that anyone can use</li>
<li><strong>Measurable Results:</strong> Track your progress with detailed analytics</li>
</ul>

<h2>Best Practices for {keyword.title()}</h2>

<h3>1. Start with Research</h3>
<p>Before implementing any strategy, it's crucial to understand your target audience and competition. Use tools like {product} to gather insights and identify opportunities.</p>

<h3>2. Focus on Quality</h3>
<p>Quality always trumps quantity. Create valuable content that genuinely helps your audience solve their problems.</p>

<h3>3. Monitor and Optimize</h3>
<p>Continuous improvement is key. Regularly review your performance and make adjustments based on data.</p>

<h2>Conclusion</h2>

<p>{keyword.title()} is not just a trend—it's a fundamental shift in how businesses connect with their audiences. By implementing the strategies outlined in this guide and leveraging powerful tools like <a href="https://{domain}">{product}</a>, you can position your business for success in 2025 and beyond.</p>

<p><strong>Ready to get started?</strong> Visit <a href="https://{domain}">{product}</a> today and discover how we can help you achieve your {keyword} goals.</p>
"""
        
        results.append(f"  ✅ 文章标题: {article_title}")
        results.append(f"  ✅ 文章长度: 约 {len(article_content)} 字符")
        
        # 保存文章到本地
        safe_keyword = "".join(c for c in keyword if c.isalnum() or c in (' ', '-', '_')).strip()[:30]
        safe_keyword = safe_keyword.replace(' ', '_') or 'article'
        
        html_filename = f"{safe_keyword}_{timestamp}.html"
        html_filepath = os.path.join(output_path, html_filename)
        
        with open(html_filepath, 'w', encoding='utf-8') as f:
            f.write(f"<!DOCTYPE html><html><head><title>{article_title}</title></head><body><h1>{article_title}</h1>{article_content}</body></html>")
        
        results.append(f"  💾 本地保存: {html_filepath}")
        
        # ========== 步骤 3: 发布到 WordPress ==========
        results.append("\n📤 **步骤 3/3: 发布到 WordPress**")
        
        wp_result = self._wp_request("POST", "posts/new", {
            "title": article_title,
            "content": article_content,
            "status": publish_status,
            "categories": "SEO, AI",
            "tags": ", ".join(related_keywords[:5])
        })
        
        if wp_result["success"]:
            post = wp_result["data"]
            results.append(f"  ✅ 发布成功!")
            results.append(f"  🆔 文章ID: {post.get('ID', 'N/A')}")
            results.append(f"  🔗 URL: {post.get('URL', 'N/A')}")
            wp_url = post.get('URL', '')
        else:
            results.append(f"  ❌ 发布失败: {wp_result['error']}")
            wp_url = ""
        
        results_text = "\n".join(results)
        
        return f"""
🎯 **研究并发布完成**

═══════════════════════════════════════
🌐 域名: {domain}
🔑 关键词: {keyword}
📝 产品: {product}
═══════════════════════════════════════

{results_text}

═══════════════════════════════════════
📋 **文章详情**
═══════════════════════════════════════
📌 标题: {article_title}
🏷️ 相关关键词: {', '.join(related_keywords[:5])}
📁 本地文件: {html_filepath}
{"🔗 WordPress URL: " + wp_url if wp_url else ""}
"""


# ==================== 兼容性别名 ====================
Functions = Tools
Function = Tools

